"""
Quotes Blueprint — 报价单相关 API 路由
从 app.py 拆分出的所有 /api/quotes/* 路由及 /api/download-logs
"""

import io
import re
from datetime import datetime
from pathlib import Path

from flask import Blueprint, request, jsonify, g, send_file
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter

from extensions import db
from models import Product, Quote, QuoteItem, User, DownloadLog
from auth import require_auth

# ─── Blueprint 定义 ──────────────────────────────────────────
quotes_bp = Blueprint('quotes', __name__, url_prefix='/api/quotes')
download_logs_bp = Blueprint('download_logs', __name__)

# 项目根目录（用于文件操作）
BASE_DIR = Path(__file__).parent
EXPORT_DIR = BASE_DIR / 'exports'
EXPORT_DIR.mkdir(exist_ok=True)


# ─── Lazy imports（避免循环依赖） ─────────────────────────────

def _check_quote_owner(quote_id):
    """指向 app.py 的 check_quote_owner"""
    from app import check_quote_owner
    return check_quote_owner(quote_id)


def _preload_products_for_quote(quote):
    """指向 app.py 的 preload_products_for_quote"""
    from app import preload_products_for_quote
    return preload_products_for_quote(quote)


def _get_setting(key, default=''):
    """读取单个系统设置（指向 app.py 的实现）"""
    from app import get_setting
    return get_setting(key, default)


# ─── 中文大写金额 ──────────────────────────────────────────

CN_NUM = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
CN_UNIT = ['', '拾', '佰', '仟']
CN_BIG_UNIT = ['', '万', '亿', '万亿']


def number_to_cn(num):
    """数字转中文大写金额"""
    if num == 0:
        return '零圆整'
    # 只处理整数部分
    num = int(num)
    if num < 0:
        return '负数'

    def _section(n):
        result = ''
        for i in range(4):
            digit = n % 10
            if digit != 0:
                result = CN_NUM[digit] + CN_UNIT[i] + result
            else:
                if result and result[0] != '零':
                    result = '零' + result
            n //= 10
        # 去掉开头多余的零
        while result.startswith('零'):
            result = result[1:]
        return result

    if num == 0:
        return '零圆整'

    result = ''
    unit_idx = 0
    while num > 0:
        section = num % 10000
        if section != 0:
            section_str = _section(section)
            if unit_idx > 0 and section < 1000:
                section_str = '零' + section_str
            result = section_str + CN_BIG_UNIT[unit_idx] + result
        elif result and result[0] not in ('零', '万', '亿'):
            # 中间有零
            pass
        num //= 10000
        unit_idx += 1

    # 处理连续的零
    while '零零' in result:
        result = result.replace('零零', '零')
    if result.endswith('零'):
        result = result[:-1]

    return result + '圆整'


# ─── Excel 公共样式 ──────────────────────────────────────────

def _excel_common_styles():
    """报价单Excel公共样式定义（export_quote_excel和_build_excel共用）"""
    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    title_font = Font(name='微软雅黑', size=10, bold=True)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    data_font = Font(name='微软雅黑', size=11, bold=True)
    total_font = Font(name='微软雅黑', size=10, bold=True)
    note_font = Font(name='微软雅黑', size=10, bold=False)
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    money_fmt = '#,##0.00'
    col_widths = [9.66, 27.16, 18.83, 20.16, 60.16, 13.33, 7.5, 11.33, 6.5, 12.16, 18.16, 16.0]
    headers = ['序号', '名称', '规格型号', '型号', '功能描述', '单价', '数量', '合计', '折扣率', '成交价', '备注', '图片']
    return {
        'YELLOW_FILL': YELLOW_FILL, 'title_font': title_font, 'header_font': header_font,
        'data_font': data_font, 'total_font': total_font, 'note_font': note_font,
        'thin_border': thin_border, 'ca': ca, 'money_fmt': money_fmt,
        'col_widths': col_widths, 'headers': headers, 'COL_COUNT': len(headers),
    }


def _excel_write_header(ws, quote, styles):
    """写入Excel前3行（公司信息+黄色标题+列表头），返回当前行号"""
    s = styles
    COL_COUNT = s['COL_COUNT']
    for ci, w in enumerate(s['col_widths'], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # 第1行：公司名 + 客户信息
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
    company = _get_setting('company_name', '').strip()
    parts = [f'公司：{company}'] if company else []
    if quote.client: parts.append(f'客户：{quote.client}')
    if quote.contact: parts.append(f'联系人：{quote.contact}')
    if quote.phone: parts.append(f'电话：{quote.phone}')
    if quote.tax_rate and quote.tax_rate > 0: parts.append(f'税率：{quote.tax_rate}%')
    if quote.quote_date: parts.append(f'日期：{quote.quote_date}')
    info = '  |  '.join(parts) if parts else ''
    c1 = ws.cell(row=1, column=1, value=info)
    c1.font = Font(name='微软雅黑', size=9, color='666666')
    c1.alignment = Alignment(horizontal='left', vertical='center')
    for ci in range(1, COL_COUNT + 1):
        ws.cell(row=1, column=ci).border = s['thin_border']
    ws.row_dimensions[1].height = 17
    # 第2行：黄色标题
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COL_COUNT)
    t = ws.cell(row=2, column=1, value=quote.title or '报价单')
    t.font = s['title_font']; t.fill = s['YELLOW_FILL']; t.alignment = s['ca']
    for ci in range(1, COL_COUNT + 1):
        ws.cell(row=2, column=ci).border = s['thin_border']
    ws.row_dimensions[2].height = 18
    # 第3行：表头
    HEAD = 3
    ws.row_dimensions[HEAD].height = 17
    for ci, h in enumerate(s['headers'], 1):
        cell = ws.cell(row=HEAD, column=ci, value=h)
        cell.font = s['header_font']; cell.alignment = s['ca']; cell.border = s['thin_border']
    return HEAD


def _build_excel(quote, pmap, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = quote.title or ''
    # ── 公共样式 + 表头 ──
    styles = _excel_common_styles()
    s = styles
    HEAD = _excel_write_header(ws, quote, styles)
    row = HEAD
    for i, item in enumerate(quote.items, 1):
        row += 1
        ws.row_dimensions[row].height = 54
        qty = item.quantity if item.quantity else 1
        up = item.unit_price if item.unit_price else 0
        subtotal = round(qty * up, 2)
        product_function_desc = ''; image_url = ''
        if item.product_id:
            product = pmap.get(item.product_id)
            if product:
                product_function_desc = product.function_desc or ''
                image_url = product.image_url
        vals = [i, item.product_name, item.product_spec or '', item.product_spec or item.product_sku or '', product_function_desc, up, qty, subtotal, 0, subtotal, item.remark or '', '']
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = s['data_font']; cell.alignment = s['ca']; cell.border = s['thin_border']
        # 嵌入产品图片到图片列（L列）— 从 BLOB 读取
        if image_url:
            try:
                img_bytes = None
                if product and hasattr(product, 'image_data') and product.image_data:
                    img_bytes = product.image_data
                if not img_bytes:
                    img_path = BASE_DIR / image_url.lstrip('/')
                    if img_path.exists():
                        img_bytes = img_path.read_bytes()
                if img_bytes:
                    img = XLImage(io.BytesIO(img_bytes))
                    w, h = img.width, img.height
                    max_w, max_h = 80, 48
                    ratio = min(max_w / w, max_h / h, 1)
                    img.width = int(w * ratio)
                    img.height = int(h * ratio)
                    col_l = get_column_letter(12)
                    col_w_px = (ws.column_dimensions[col_l].width or 10) * 7
                    row_h_pt = ws.row_dimensions[row].height or 60
                    x_emu = int(max(0, (col_w_px - img.width) / 2) * 9525)
                    y_emu = int(max(0, (row_h_pt - img.height) / 2) * 9525)
                    img.anchor = TwoCellAnchor(
                        _from=AnchorMarker(col=11, colOff=x_emu, row=row-1, rowOff=y_emu),
                        to=AnchorMarker(col=11, colOff=x_emu + img.width * 9525, row=row-1, rowOff=y_emu + img.height * 9525)
                    )
                    ws.add_image(img)
            except Exception:
                pass
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    total_amt = quote.total_amount or 0
    tlabel = ws.cell(row=row, column=1, value=f'合计（大写）：{number_to_cn(total_amt)}')
    tlabel.font = s['total_font']; tlabel.alignment = Alignment(horizontal='right', vertical='center'); tlabel.border = s['thin_border']
    for ci in range(2, 11):
        c = ws.cell(row=row, column=ci); c.font = s['total_font']; c.border = s['thin_border']
    tc = ws.cell(row=row, column=11, value=total_amt)
    tc.font = s['total_font']; tc.number_format = s['money_fmt']; tc.alignment = s['ca']; tc.border = s['thin_border']
    ws.cell(row=row, column=12).border = s['thin_border']; ws.cell(row=row, column=12).font = s['total_font']
    row += 1
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=s['COL_COUNT'])
    nc = ws.cell(row=row, column=1, value=quote.remark or '注：硬件默认自验收日起维保1年，硬件1年内享受免费寄修服务。')
    nc.font = s['note_font']; nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for ci in range(1, s['COL_COUNT'] + 1):
        ws.cell(row=row, column=ci).border = s['thin_border']
    footer = _get_setting('footer_text', '').strip()
    if footer:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=s['COL_COUNT'])
        fc = ws.cell(row=row, column=1, value=footer)
        fc.font = Font(name='微软雅黑', size=9, color='888888')
        fc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30
        for ci in range(1, s['COL_COUNT'] + 1):
            ws.cell(row=row, column=ci).border = Border()
    wb.save(filepath)


# ─── Quotes API 路由 ──────────────────────────────────────

@quotes_bp.route('', methods=['GET'])
def list_quotes():
    """报价单列表，支持分页、状态筛选、关键词搜索（含拼音）"""
    import re
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 200)
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()

    query = Quote.query
    # 非管理员只看自己的报价单
    if hasattr(g, 'current_user') and g.current_user and g.current_user.role != 'admin':
        query = query.filter(Quote.created_by == g.current_user.id)
    if status_filter:
        query = query.filter(Quote.status == status_filter)

    # 拼音搜索：纯ASCII（无汉字）时启用
    is_pinyin = search and not re.search(r'[\u4e00-\u9fff]', search)
    if is_pinyin:
        from pypinyin import pinyin, Style
        q_lower = search.lower().strip()
        all_quotes = query.order_by(Quote.id.desc()).all()

        def pinyin_match(q):
            texts = [q.title or '', q.client or '']
            for text in texts:
                if not text:
                    continue
                py_list = pinyin(text, style=Style.NORMAL, heteronym=False)
                full_py = ''.join(p[0] for p in py_list).lower()
                if q_lower in full_py:
                    return True
                initials = ''.join(p[0][0] for p in py_list).lower()
                if q_lower in initials:
                    return True
                if len(q_lower) >= 2 and len(initials) >= 2:
                    if q_lower in initials:
                        return True
            return False

        filtered = [q for q in all_quotes if pinyin_match(q)]
        total = len(filtered)
        quotes = filtered[(page - 1) * per_page: page * per_page]
    else:
        query = query.order_by(Quote.id.desc())
        if search:
            like = f'%{search}%'
            query = query.filter(
                db.or_(Quote.title.ilike(like), Quote.client.ilike(like))
            )
        total = query.count()
        quotes = query.offset((page - 1) * per_page).limit(per_page).all()

    # 预加载所有创建者用户名，避免 N+1 查询
    creator_ids = list(set(q.created_by for q in quotes if q.created_by))
    users_map = {}
    if creator_ids:
        users = User.query.filter(User.id.in_(creator_ids)).all()
        users_map = {u.id: u.username for u in users}

    return jsonify({
        'quotes': [q.to_dict(users_map=users_map) for q in quotes],
        'total': total,
        'page': page,
        'per_page': per_page,
    })


@quotes_bp.route('/stats', methods=['GET'])
def quote_stats():
    """按客户统计报价单（客户维度聚合）"""
    qf = Quote.client.isnot(None), Quote.client != ''
    if hasattr(g, 'current_user') and g.current_user and g.current_user.role != 'admin':
        qf = qf + (Quote.created_by == g.current_user.id,)
    rows = db.session.query(
        Quote.client, Quote.id, Quote.title, Quote.total_amount,
        Quote.status, Quote.quote_date, Quote.download_count
    ).filter(*qf)\
     .order_by(Quote.client, Quote.id.desc()).all()

    customers = {}
    for client, qid, title, amt, status, qdate, dl in rows:
        if client not in customers:
            customers[client] = {'client': client, 'quotes': [], 'total_amount': 0, 'quote_count': 0}
        customers[client]['quotes'].append({
            'id': qid, 'title': title, 'total_amount': amt or 0,
            'status': status, 'quote_date': qdate, 'download_count': dl or 0
        })
        customers[client]['total_amount'] += (amt or 0)
        customers[client]['quote_count'] += 1

    return jsonify({'customers': sorted(customers.values(), key=lambda x: x['total_amount'], reverse=True)})


@quotes_bp.route('/<int:quote_id>/status', methods=['PATCH'])
def update_quote_status(quote_id):
    """修改报价单状态"""
    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status
    data = request.get_json()
    new_status = data.get('status', '')
    valid_statuses = ['draft', 'sent', 'confirmed', 'rejected', 'expired']
    if new_status not in valid_statuses:
        return jsonify({'error': f'无效状态，可选: {valid_statuses}'}), 400
    quote.status = new_status
    db.session.commit()
    return jsonify({'quote': quote.to_dict()})


@quotes_bp.route('', methods=['POST'])
def create_quote():
    data = request.get_json()
    if not data:
        return jsonify({'error': '缺少数据'}), 400

    quote = Quote(
        title=data.get('title', ''),
        client=data.get('client', ''),
        contact=data.get('contact', ''),
        phone=data.get('phone', ''),
        quote_date=data.get('quote_date', datetime.now().strftime('%Y-%m-%d')),
        valid_days=int(data.get('valid_days', 15)),
        tax_rate=round(float(data.get('tax_rate', 0)), 2),
        remark=data.get('remark', ''),
        created_by=g.current_user.id if hasattr(g, 'current_user') and g.current_user else None,
    )

    items_data = data.get('items', [])
    total = 0
    # 预加载产品信息以填充 name/spec/unit/sku
    pids = [it.get('product_id') for it in items_data if it.get('product_id')]
    pmap = {}
    if pids:
        products = Product.query.filter(Product.id.in_(pids)).all()
        pmap = {p.id: p for p in products}
    for i, item in enumerate(items_data):
        qty = int(item.get('quantity', 1))
        up = round(float(item.get('unit_price', 0)), 2)
        amt = round(qty * up, 2)
        pid = item.get('product_id')
        prod = pmap.get(pid) if pid else None
        qi = QuoteItem(
            product_id=pid,
            product_name=item.get('product_name') or (prod.name if prod else ''),
            product_sku=item.get('product_sku') or (prod.sku if prod else ''),
            product_spec=item.get('product_spec') or (prod.spec if prod else ''),
            product_unit=item.get('product_unit') or (prod.unit if prod else ''),
            quantity=qty,
            unit_price=up,
            amount=amt,
            remark=item.get('remark', ''),
            sort_order=i,
        )
        quote.items.append(qi)
        total += amt

    quote.total_amount = round(total, 2)
    db.session.add(quote)
    db.session.commit()
    return jsonify({'quote': quote.to_dict()}), 201


@quotes_bp.route('/<int:quote_id>', methods=['GET'])
def get_quote(quote_id):
    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status
    pmap = _preload_products_for_quote(quote)
    return jsonify({'quote': quote.to_dict(pmap)})


@quotes_bp.route('/<int:quote_id>', methods=['PUT'])
def update_quote(quote_id):
    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status
    data = request.get_json()
    if data.get('title') is not None: quote.title = data['title']
    if data.get('client') is not None: quote.client = data['client']
    if data.get('contact') is not None: quote.contact = data['contact']
    if data.get('phone') is not None: quote.phone = data['phone']
    if data.get('quote_date') is not None: quote.quote_date = data['quote_date']
    if data.get('valid_days') is not None: quote.valid_days = int(data['valid_days'])
    if data.get('tax_rate') is not None: quote.tax_rate = round(float(data['tax_rate']), 2)
    if data.get('remark') is not None: quote.remark = data['remark']
    if data.get('status') is not None: quote.status = data['status']

    if 'items' in data:
        QuoteItem.query.filter_by(quote_id=quote_id).delete()
        total = 0
        # 预加载产品信息以填充 name/spec/unit/sku
        pids = [it.get('product_id') for it in data['items'] if it.get('product_id')]
        pmap = {}
        if pids:
            products = Product.query.filter(Product.id.in_(pids)).all()
            pmap = {p.id: p for p in products}
        for i, item in enumerate(data['items']):
            qty = int(item.get('quantity', 1))
            up = round(float(item.get('unit_price', 0)), 2)
            amt = round(qty * up, 2)
            pid = item.get('product_id')
            prod = pmap.get(pid) if pid else None
            qi = QuoteItem(
                quote_id=quote_id,
                product_id=pid,
                product_name=item.get('product_name') or (prod.name if prod else ''),
                product_sku=item.get('product_sku') or (prod.sku if prod else ''),
                product_spec=item.get('product_spec') or (prod.spec if prod else ''),
                product_unit=item.get('product_unit') or (prod.unit if prod else ''),
                quantity=qty,
                unit_price=up,
                amount=amt,
                remark=item.get('remark', ''),
                sort_order=i,
            )
            db.session.add(qi)
            total += amt
        quote.total_amount = round(total, 2)

    db.session.commit()
    return jsonify({'quote': quote.to_dict()})


@quotes_bp.route('/<int:quote_id>', methods=['DELETE'])
def delete_quote(quote_id):
    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status
    db.session.delete(quote)
    db.session.commit()
    return jsonify({'message': '已删除'})


@quotes_bp.route('/batch', methods=['DELETE'])
@require_auth
def batch_delete_quotes():
    """批量删除报价单（仅限自己创建的或管理员删除全部）"""
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': '请提供要删除的报价单 ID 列表'}), 400
    if len(ids) > 100:
        return jsonify({'error': '单次最多删除 100 条'}), 400

    user = g.current_user
    is_admin = user.role == 'admin'

    quotes = Quote.query.filter(Quote.id.in_(ids)).all()
    deletable = []
    forbidden = []
    for q in quotes:
        if is_admin or q.created_by == user.id:
            deletable.append(q)
        else:
            forbidden.append(q.id)

    for q in deletable:
        db.session.delete(q)
    db.session.commit()

    return jsonify({
        'deleted': len(deletable),
        'total': len(ids),
        'forbidden': forbidden,
    })


@quotes_bp.route('/<int:quote_id>/export-excel', methods=['GET'])
def export_quote_excel(quote_id):
    """导出报价单 — 样式精确克隆模板.xlsx"""
    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status

    # 记录下载次数
    quote.download_count = (quote.download_count or 0) + 1
    db.session.commit()

    pmap = _preload_products_for_quote(quote)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = quote.title or '报价单'

    # ── 公共样式 + 表头 ──
    styles = _excel_common_styles()
    HEAD = _excel_write_header(ws, quote, styles)
    s = styles
    pct_fmt = '0%'

    # ── 数据行 ──
    row = HEAD
    for i, item in enumerate(quote.items, 1):
        row += 1
        ws.row_dimensions[row].height = 54

        qty = item.quantity if item.quantity else 1
        up = item.unit_price if item.unit_price else 0
        subtotal = round(qty * up, 2)

        # 取产品 function_desc 作为功能描述
        product_function_desc = ''
        image_url = None
        if item.product_id:
            product = pmap.get(item.product_id)
            if product:
                product_function_desc = product.function_desc or ''
                image_url = product.image_url

        desc = product_function_desc

        vals = [i, item.product_name, item.product_spec or '',
                item.product_spec or item.product_sku or '', desc,
                up, qty, subtotal, 0, subtotal, item.remark or '', '']

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = s['data_font']; cell.alignment = s['ca']; cell.border = s['thin_border']
            if ci in (6, 8, 10): cell.number_format = s['money_fmt']
            elif ci == 9: cell.number_format = pct_fmt

        ws.cell(row=row, column=1).border = s['thin_border']
        ws.cell(row=row, column=s['COL_COUNT']).border = s['thin_border']

        # 嵌入产品图片到图片列（L列）— 从 BLOB 读取
        if image_url:
            try:
                img_bytes = None
                if product and hasattr(product, 'image_data') and product.image_data:
                    img_bytes = product.image_data
                if not img_bytes:
                    img_path = BASE_DIR / image_url.lstrip('/')
                    if img_path.exists():
                        img_bytes = img_path.read_bytes()
                if img_bytes:
                    img = XLImage(io.BytesIO(img_bytes))
                    # 限制尺寸适配图片列：宽≈80px, 高≤48px
                    w, h = img.width, img.height
                    max_w, max_h = 80, 48
                    ratio = min(max_w / w, max_h / h, 1)
                    img.width = int(w * ratio)
                    img.height = int(h * ratio)
                    # 图片单元格内居中
                    col_l = get_column_letter(12)
                    col_w_px = (ws.column_dimensions[col_l].width or 10) * 7
                    row_h_pt = ws.row_dimensions[row].height or 60
                    x_emu = int(max(0, (col_w_px - img.width) / 2) * 9525)
                    y_emu = int(max(0, (row_h_pt - img.height) / 2) * 9525)
                    img.anchor = TwoCellAnchor(
                        _from=AnchorMarker(col=11, colOff=x_emu, row=row-1, rowOff=y_emu),
                        to=AnchorMarker(col=11, colOff=x_emu + img.width * 9525, row=row-1, rowOff=y_emu + img.height * 9525)
                    )
                    ws.add_image(img)
            except Exception:
                pass

    # ── 合计行 ──
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    total_amt = quote.total_amount or 0
    tlabel = ws.cell(row=row, column=1, value=f'合计（大写）：{number_to_cn(total_amt)}')
    tlabel.font = s['total_font']
    tlabel.alignment = Alignment(horizontal='right', vertical='center')
    tlabel.border = s['thin_border']

    for ci in range(2, 11):
        c = ws.cell(row=row, column=ci)
        c.font = s['total_font']; c.border = s['thin_border']

    tc = ws.cell(row=row, column=11, value=total_amt)
    tc.font = s['total_font']; tc.number_format = s['money_fmt']; tc.alignment = s['ca']
    tc.border = s['thin_border']

    ws.cell(row=row, column=12).border = s['thin_border']
    ws.cell(row=row, column=12).font = s['total_font']

    # ── 备注行 ──
    row += 1
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=s['COL_COUNT'])
    nc = ws.cell(row=row, column=1, value=quote.remark or '注：硬件默认自验收日起维保1年，硬件1年内享受免费寄修服务。')
    nc.font = s['note_font']
    nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for ci in range(1, s['COL_COUNT'] + 1):
        ws.cell(row=row, column=ci).border = s['thin_border']

    # ── 页脚行（公司自定义） ──
    footer = _get_setting('footer_text', '').strip()
    if footer:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=s['COL_COUNT'])
        fc = ws.cell(row=row, column=1, value=footer)
        fc.font = Font(name='微软雅黑', size=9, color='888888')
        fc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30
        for ci in range(1, s['COL_COUNT'] + 1):
            ws.cell(row=row, column=ci).border = Border()

    # ── 打印：纵向 ──
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.4; ws.page_margins.right = 0.4

    filepath = EXPORT_DIR / f'报价单_{quote.id}.xlsx'
    wb.save(filepath)
    # 优先使用前端传来的浏览器本地日期，兜底用服务器日期
    download_date = request.args.get('download_date', '').strip()
    if download_date:
        date_str = download_date
    else:
        date_str = datetime.now().strftime('%Y%m%d')
    dl_name = f'{quote.client or ""}_{quote.title or ""}_{quote.contact or ""}_{date_str}'.strip('_').replace(' ','') + '.xlsx'

    # 记录下载日志
    user_name = g.current_user.username if hasattr(g, 'current_user') and g.current_user else request.args.get('user_name', '').strip()
    if user_name:
        log = DownloadLog(quote_id=quote_id, user_name=user_name)
        db.session.add(log)
        db.session.commit()

    return send_file(str(filepath), download_name=dl_name, as_attachment=True)


# ─── 邮件发送 (v1.4.0) ───
@quotes_bp.route('/<int:quote_id>/send-email', methods=['POST'])
def send_quote_email(quote_id):
    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status
    data = request.get_json(silent=True) or {}
    to_email = data.get('to_email', '').strip()
    if not to_email:
        return jsonify({'error': '请填写收件人邮箱'}), 400
    subject = data.get('subject', '').strip()
    body_text = data.get('body', '').strip()
    smtp_host = _get_setting('smtp_host', '')
    if not smtp_host:
        return jsonify({'error': 'SMTP未配置'}), 400
    smtp_port = int(_get_setting('smtp_port', '587'))
    smtp_user = _get_setting('smtp_user', '')
    smtp_password = _get_setting('smtp_password', '')
    smtp_from = _get_setting('smtp_from', smtp_user)
    smtp_use_tls = _get_setting('smtp_use_tls', 'true').lower() == 'true'

    # 生成附件
    filepath = EXPORT_DIR / f'报价单_{quote_id}.xlsx'
    pmap = _preload_products_for_quote(quote)
    _build_excel(quote, pmap, str(filepath))

    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    msg = MIMEMultipart()
    msg['From'] = smtp_from
    msg['To'] = to_email
    msg['Subject'] = subject or f'{quote.title or quote.client or ""}'
    msg.attach(MIMEText(body_text or f'{quote.title or ""}\n{quote.client or ""}\n{quote.quote_date or ""}', 'plain', 'utf-8'))
    with open(filepath, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=f'{quote.client or ""}_{quote.title or ""}.xlsx')
        msg.attach(part)
    try:
        if smtp_use_tls:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=15)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=15)
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        server.quit()
        return jsonify({'success': True, 'message': f'{to_email}'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── 下载日志 API ───
@download_logs_bp.route('/api/download-logs', methods=['GET'])
def list_download_logs():
    logs = DownloadLog.query.order_by(DownloadLog.downloaded_at.desc()).limit(200).all()
    return jsonify({'logs': [log.to_dict() for log in logs]})


@download_logs_bp.route('/api/download-logs/stats', methods=['GET'])
def download_logs_stats():
    """按用户汇总下载次数"""
    rows = db.session.query(
        DownloadLog.user_name, func.count(DownloadLog.id)
    ).group_by(DownloadLog.user_name).order_by(func.count(DownloadLog.id).desc()).all()
    return jsonify({'users': [{'user_name': name, 'count': cnt} for name, cnt in rows]})


# ─── 报价单 HTML 预览 ───
@quotes_bp.route('/<int:quote_id>/preview', methods=['GET'])
def preview_quote_html(quote_id):
    """返回报价单的HTML预览（17列格式匹配原模板）"""
    import html as _html  # 用于转义用户输入防XSS

    quote, err, status = _check_quote_owner(quote_id)
    if not quote:
        return err, status

    pmap = _preload_products_for_quote(quote)

    def fmt(n):
        if n is None: return '0.00'
        return f'{n:,.2f}'

    def fmt_int(n):
        if n is None: return ''
        try: return f'{int(float(n))}'
        except Exception: return str(n)

    def e(s):
        """html.escape shorthand — 防止用户输入中的HTML/JS注入"""
        return _html.escape(str(s)) if s else ''

    info_parts = []
    if quote.client: info_parts.append(f'客户：{e(quote.client)}')
    if quote.contact: info_parts.append(f'联系人：{e(quote.contact)}')
    if quote.phone: info_parts.append(f'电话：{e(quote.phone)}')
    if quote.quote_date: info_parts.append(f'日期：{e(quote.quote_date)}')
    if quote.valid_days: info_parts.append(f'有效期：{quote.valid_days}天')
    if quote.tax_rate and quote.tax_rate > 0: info_parts.append(f'税率：{quote.tax_rate}%')
    info_line = '  |  '.join(info_parts) if info_parts else ''

    items_html = ''
    for i, item in enumerate(quote.items, 1):
        supplier = ''; supplier_sku = ''; cost = 0; prod_function_desc = ''; image_url = ''
        if item.product_id:
            prod = pmap.get(item.product_id)
            if prod:
                supplier = prod.supplier or ''
                supplier_sku = prod.spec or prod.sku or ''
                cost = prod.cost_price or 0
                prod_function_desc = prod.function_desc or ''
                image_url = prod.image_url or ''

        qty = item.quantity if item.quantity else 1
        up = item.unit_price if item.unit_price else 0
        subtotal = round(qty * up, 2)
        deal_price = subtotal
        guide_price = round(cost * 1.5, 2) if cost else 0
        min_retail = round(cost * 1.15, 2) if cost else 0

        # 图片列：使用 /api/products/<id>/image 端点（附带 token 供 img 标签鉴权）
        img_cell = ''
        if image_url:
            _preview_token = request.headers.get('Authorization', '').replace('Bearer ', '')
            src = f'/quote/api/products/{item.product_id}/image?token={_preview_token}'
            img_cell = f'<img src="{src}" style="max-width:100px;max-height:48px;object-fit:contain;display:block;margin:0 auto">'
        else:
            img_cell = '—'

        items_html += f'''
        <tr>
            <td>{i}</td>
            <td>{e(item.product_name)}</td>
            <td>{e(item.product_spec or '')}</td>
            <td>{e(item.product_sku or supplier_sku)}</td>
            <td>{e(prod_function_desc or '')}</td>
            <td>{fmt(up)}</td>
            <td>{fmt_int(qty)}</td>
            <td>{fmt(subtotal)}</td>
            <td>0%</td>
            <td>{fmt(deal_price)}</td>
            <td>{e(item.remark or '')}</td>
            <td style="text-align:center;vertical-align:middle">{img_cell}</td>
        </tr>'''

    html = f'''<style>
.pv-table{{width:100%;border-collapse:collapse;font-size:11pt;font-weight:bold}}
.pv-table th{{font-size:10pt;font-weight:bold;padding:3px 2px;border:1px solid #ccc;text-align:center;background:#fff}}
.pv-table td{{padding:3px 2px;border:1px solid #ccc;vertical-align:middle;text-align:center}}
.pv-table td:first-child{{border-left:1px solid #ccc}}
.pv-table td:last-child{{border-right:1px solid #ccc}}
.pv-table tr:hover td{{background:#fffbe6}}
.pv-table .total-row td{{font-size:10pt;font-weight:bold;border-top:1px solid #ccc;border-bottom:1px solid #ccc;padding:4px 2px;background:#fafafa}}
.pv-table .total-row td:first-child{{border-left:1px solid #ccc}}
.pv-table .total-row td:last-child{{border-right:1px solid #ccc}}
.pv-table .total-amount{{font-size:10pt}}
.pv-note{{font-size:10pt;padding:3px 8px;border:1px solid #ccc;border-top:none}}
</style>
<div style="overflow-x:auto">
<table class="pv-table">
  <thead>
    <tr>
      <td colspan="12" style="font-size:9pt;color:#666;padding:4px 8px;text-align:left;font-weight:normal">{info_line}</td>
    </tr>
    <tr>
      <th colspan="12" style="background:#FFFF00;font-size:10pt;font-weight:bold;text-align:center;padding:4px">{e(quote.title or '报价单')}</th>
    </tr>
    <tr>
      <th style="width:50px">序号</th>
      <th style="width:170px">名称</th>
      <th style="width:100px">规格型号</th>
      <th style="width:110px">型号</th>
      <th style="width:300px">功能描述</th>
      <th style="width:75px">单价</th>
      <th style="width:45px">数量</th>
      <th style="width:70px">合计</th>
      <th style="width:42px">折扣率</th>
      <th style="width:75px">成交价</th>
      <th style="width:90px">备注</th>
      <th style="width:80px">图片</th>
    </tr>
  </thead>
  <tbody>
    {items_html}
  </tbody>
  <tfoot>
    <tr class="total-row">
      <td colspan="11" style="text-align:right">合计（大写）：<strong>{number_to_cn(quote.total_amount or 0)}</strong></td>
      <td class="total-amount">¥{fmt(quote.total_amount or 0)}</td>
    </tr>
  </tfoot>
</table>
</div>
<div class="pv-note">{e(quote.remark or '注：硬件默认自验收日起维保1年，硬件1年内享受免费寄修服务。')}</div>'''
    return html
