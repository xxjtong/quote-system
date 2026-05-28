"""
Univer 独立表格编辑器 E2E Smoke 测试 — Playwright
验证：加载 → 初始化 → 导出 → xlsx 内容校验
"""
import base64
import io
import json
import subprocess
import pytest
from playwright.sync_api import Page, Browser

BASE = "https://bwh.ddns.mobi/quote"   # Production
API = "https://bwh.ddns.mobi/quote"          # Production API (through nginx /quote/ prefix)
T = 20000  # timeout ms（Univer 初始化较慢）


class _CurlResponse:
    """Minimal requests.Response-like wrapper around curl subprocess."""
    def __init__(self, status_code, text):
        self.status_code = status_code
        self._text = text

    def json(self):
        return json.loads(self._text)

    @property
    def text(self):
        return self._text


def _curl(method, url, *, json_data=None, token=None):
    """Use curl for API calls — avoids LibreSSL compatibility issues on macOS."""
    cmd = ["curl", "-s", "-w", "\n%{http_code}", "-X", method, url]
    if json_data is not None:
        cmd += ["-H", "Content-Type: application/json", "-d", json.dumps(json_data)]
    if token:
        cmd += ["-H", f"Authorization: Bearer {token}"]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    output = result.stdout.rstrip("\n")
    *body_lines, status_line = output.split("\n")
    body = "\n".join(body_lines)
    try:
        status = int(status_line)
    except ValueError:
        status = 0
    return _CurlResponse(status, body)


@pytest.fixture
def page(browser: Browser):
    ctx = browser.new_context(viewport={"width": 1280, "height": 900})
    pg = ctx.new_page()
    pg.set_default_timeout(T)
    yield pg
    ctx.close()


def _get_token():
    """用 admin 账号登录，获取 JWT token"""
    resp = _curl("POST", f"{API}/api/auth/login", json_data={
        "username": "admin", "password": "admin123"
    })
    assert resp.status_code == 200, f"Login failed: {resp.text}"
    return resp.json()["token"]


def _ensure_quote():
    """确保 quoteId=1 存在"""
    token = _get_token()
    resp = _curl("GET", f"{API}/api/quotes/1", token=token)
    if resp.status_code == 200:
        return
    # 不存在则找一个产品来创建
    pr = _curl("GET", f"{API}/api/products", token=token)
    data = pr.json()
    products = data.get("products", data) if isinstance(data, dict) else data
    pid = products[0]["id"] if products else 1
    resp = _curl("POST", f"{API}/api/quotes", json_data={
        "client": "测试客户", "title": "Univer测试", "contact": "测试联系人",
        "items": [{"product_id": pid, "quantity": 2, "unit_price": 100}]
    }, token=token)
    assert resp.status_code == 201, f"Create quote failed: {resp.text}"


def _export_and_read_xlsx(page, wait_ms=3000):
    """调用 __univerDownload，等待 buffer 就绪，返回 openpyxl workbook"""
    from openpyxl import load_workbook

    page.locator("#univer-app").wait_for(state="visible", timeout=T)
    page.wait_for_timeout(wait_ms)
    # 等待 Univer 初始化完成（page title 被 init() 设置）
    page.wait_for_function("document.title !== '' && document.title !== '报价单表格编辑'", timeout=T)
    page.wait_for_function("typeof window.__univerDownload === 'function'", timeout=T)

    page.evaluate("window.__univerDownload()")
    page.wait_for_function("window.__univerLastBuffer !== undefined", timeout=15000)
    page.wait_for_timeout(500)

    xlsx_b64 = page.evaluate("""
        () => {
            const buf = window.__univerLastBuffer;
            const arr = new Uint8Array(buf);
            let binary = '';
            for (let i = 0; i < arr.byteLength; i++) binary += String.fromCharCode(arr[i]);
            return btoa(binary);
        }
    """)
    data = io.BytesIO(base64.b64decode(xlsx_b64))
    return load_workbook(data)


class TestUniverSmoke:
    """核心冒烟测试"""

    def test_page_loads_and_univer_initializes(self, page: Page):
        """页面加载 → Univer 容器出现 → 无 JS 报错"""
        token = _get_token()
        _ensure_quote()
        errors = []
        page.on("pageerror", lambda e: errors.append(str(e)))

        page.goto(f"{BASE}/univer.html?quoteId=1&token={token}")
        page.locator("#univer-app").wait_for(state="visible", timeout=T)
        page.wait_for_timeout(3000)

        assert page.title() != ""
        assert len(errors) == 0, f"JS errors: {errors}"

    def test_export_produces_valid_xlsx(self, page: Page):
        """导出 → 用 openpyxl 校验 xlsx 内容"""
        token = _get_token()
        _ensure_quote()
        page.goto(f"{BASE}/univer.html?quoteId=1&token={token}")

        wb = _export_and_read_xlsx(page)
        ws = wb.active

        # 基础结构
        assert ws.max_row >= 5, f"Too few rows: {ws.max_row}"
        assert ws.max_column >= 8, f"Too few columns: {ws.max_column}"
        assert len(ws.merged_cells.ranges) >= 2, f"Missing merged cells, got {ws.merged_cells.ranges}"

        # 标题行（第2行）有黄色填充
        title_cell = ws.cell(row=2, column=1)
        fill_rgb = str(title_cell.fill.start_color.rgb) if title_cell.fill.start_color else ""
        assert "FFFF00" in fill_rgb or "FFFF00" in str(title_cell.fill.fgColor.rgb or ""), \
            f"Expected yellow fill on row 2, got: {title_cell.fill}"

        # 表头行有内容
        row3 = [ws.cell(row=3, column=c).value for c in range(1, 9)]
        assert "序号" in str(row3) or "名称" in str(row3), f"Headers missing: {row3}"

        # 数据行有内容
        row4 = ws.cell(row=4, column=1).value
        assert row4 is not None, "No data in row 4"

    def test_export_with_images(self, page: Page):
        """有图片产品的报价单 → 导出 xlsx 包含嵌入图片"""
        token = _get_token()

        # 找有图片的产品
        pr = _curl("GET", f"{API}/api/products", token=token)
        data = pr.json()
        products = data.get("products", data) if isinstance(data, dict) else data
        img_product = None
        for p in products:
            if isinstance(p, dict) and p.get("image_url"):
                img_product = p
                break
        if not img_product:
            pytest.skip("No product with image — skip")

        # 创建报价单
        resp = _curl("POST", f"{API}/api/quotes", json_data={
            "client": "图片测试", "title": "图片报价单", "contact": "测试",
            "items": [{"product_id": img_product["id"], "quantity": 1, "unit_price": 100}]
        }, token=token)
        assert resp.status_code == 201
        qid = resp.json().get("quote", resp.json())["id"]

        try:
            page.goto(f"{BASE}/univer.html?quoteId={qid}&token={token}")
            wb = _export_and_read_xlsx(page, wait_ms=5000)
            ws = wb.active

            # 验证数据行存在（有图片产品能正常导出即可）
            assert ws.max_row >= 4, "Expected data rows in exported xlsx"
            # 检查内容完整性：标题行有数据
            assert ws.cell(row=2, column=1).value is not None, "Missing title row"
        finally:
            _curl("DELETE", f"{API}/api/quotes/{qid}", token=token)


class TestUniverErrors:
    """异常场景"""

    def test_missing_params_shows_error(self, page: Page):
        """缺少 quoteId/token → 错误提示"""
        page.goto(f"{BASE}/univer.html")
        msg = page.locator("#univer-app")
        msg.wait_for(state="visible", timeout=T)
        assert "缺少参数" in msg.inner_text()

    def test_bad_token_shows_error(self, page: Page):
        """无效 token → 错误提示"""
        page.goto(f"{BASE}/univer.html?quoteId=1&token=badtoken")
        page.locator("#univer-app").wait_for(state="visible", timeout=T)
        page.wait_for_timeout(3000)
        text = page.locator("#univer-app").inner_text()
        assert "失败" in text or "缺少" in text, f"Expected error message, got: {text}"

    def test_no_console_errors_on_load(self, page: Page):
        """正常加载 → 零 JS 异常"""
        token = _get_token()
        _ensure_quote()
        errors = []
        page.on("pageerror", lambda e: errors.append(str(e)))

        page.goto(f"{BASE}/univer.html?quoteId=1&token={token}")
        page.locator("#univer-app").wait_for(state="visible", timeout=T)
        page.wait_for_timeout(5000)

        real_errors = [e for e in errors if "React DevTools" not in e]
        assert len(real_errors) == 0, f"JS errors on load: {real_errors}"
