"""
Products Blueprint — 产品相关 API 路由
从 app.py 拆分出的所有 /api/products/* 路由及 /api/upload/image、/api/download-image
"""

import json
import os
import io
import re
import random
from datetime import datetime
from pathlib import Path

from flask import Blueprint, request, jsonify, g, Response, send_file
from sqlalchemy import func, distinct

from extensions import db
from models import (Product, AIUsageLog, User, Quote,
    ProductImage, ProductCommMethod, ProductCommProtocol,
    ProductPowerSupply)
from auth import require_auth, require_admin

from utils import _debug_log, _log_ai_usage, _safe_number, _compute_pinyin_search
from product_utils import (
    compress_image_if_needed, _ocr_fallback, doubao_vision_recognize,
    _parse_json_reply, _product_from_parsed, deepseek_parse_product,
    smart_parse_product, parse_product_line,
)

# ─── Blueprint 定义 ──────────────────────────────────────────
# 所有产品相关路由共用此蓝图（upload/download-image 用完整路径）
products_bp = Blueprint('products', __name__)

# 项目根目录（用于文件操作）
BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

# Gateway URL（deepseek_parse_product 用）
_gateway_url = os.environ.get('QUOTE_GATEWAY_URL', 'http://127.0.0.1:8642')


# ─── 辅助函数包装（已验证模块） ─────────────────────────────

def _get_setting(key, default=''):
    """读取单个系统设置（本地包装，指向 helpers.py 的实现）"""
    from helpers import get_setting
    return get_setting(key, default)


def _filter_fields_for_user(data_dict, is_admin):
    """过滤非管理员不可见字段（指向 helpers.py 的实现）"""
    from helpers import filter_fields_for_user
    return filter_fields_for_user(data_dict, is_admin)


def _get_field_visibility():
    """获取字段可见性（指向 helpers.py 的实现）"""
    from helpers import get_field_visibility
    return get_field_visibility()


# ─── 辅助函数 ────────────────────────────────────────────────

def _store_image_blob(product, data):
    """[DEPRECATED] 图片已通过 product_images 表和磁盘文件管理，不再写入 BLOB。
    保留空函数以兼容旧调用点。"""
    pass


def _cleanup_image_files(image_url):
    """删除磁盘上的图片文件（含缩略图）。"""
    if not image_url or not image_url.startswith('/uploads/'):
        return
    fpath = BASE_DIR / image_url.lstrip('/')
    try:
        fpath.unlink(missing_ok=True)
        # Also delete thumbnail
        stem = fpath.stem
        if not stem.endswith('_thumb'):
            thumb = fpath.parent / f'{stem}_thumb.jpg'
            thumb.unlink(missing_ok=True)
    except Exception:
        pass


def _cleanup_product_images(product_id):
    """删除产品关联的所有磁盘图片文件。"""
    from models import ProductImage as PImg
    imgs = PImg.query.filter_by(product_id=product_id).all()
    for img in imgs:
        _cleanup_image_files(img.url)


def add_pinyin_field(p_dict):
    """给产品字典添加 _py 字段（全拼+首字母），供前端拼音搜索用。"""
    from pypinyin import pinyin, Style
    texts = [p_dict.get('name',''), p_dict.get('spec',''), p_dict.get('supplier',''), p_dict.get('function_desc','')]
    py_parts = []
    initials_parts = []
    for t in texts:
        if t:
            py_list = pinyin(t, style=Style.NORMAL, heteronym=False)
            py_parts.append(''.join(p[0] for p in py_list).lower())
            initials_parts.append(''.join(p[0][0] for p in py_list).lower())
    p_dict['_py'] = ' '.join(py_parts)
    p_dict['_py_initials'] = ' '.join(initials_parts)
    return p_dict


# ─── 产品路由 ────────────────────────────────────────────────

def _parse_int_list(raw):
    """Parse comma-separated int list from query string, e.g. '1,3,5' → [1,3,5]"""
    if not raw or not raw.strip():
        return []
    result = []
    for part in raw.split(','):
        part = part.strip()
        if part:
            try:
                result.append(int(part))
            except ValueError:
                pass
    return result


@products_bp.route('/api/products', methods=['GET'])
def list_products():
    """产品列表，支持搜索（含拼音）和分类筛选"""
    import re
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)
    search = request.args.get('search', '').strip()
    category = request.args.get('category', '').strip()
    supplier = request.args.get('supplier', '').strip()
    category_id = request.args.get('category_id', type=int)
    comm_method = _parse_int_list(request.args.get('comm_method', ''))
    power_supply = _parse_int_list(request.args.get('power_supply', ''))
    sensor_metric = _parse_int_list(request.args.get('sensor_metric', ''))
    manufacturer_id = request.args.get('manufacturer_id', type=int)

    # 排序
    sort_by = request.args.get('sort_by', 'id')
    sort_order = request.args.get('sort_order', 'desc')
    if sort_by == 'name':
        col = Product.name
    elif sort_by == 'price':
        col = Product.price
    elif sort_by == 'category':
        col = Product.category
    else:
        col = Product.id

    query = Product.query
    is_admin = hasattr(g, 'current_user') and g.current_user and g.current_user.role == 'admin'
    if not is_admin:
        uid = g.current_user.id if hasattr(g, 'current_user') and g.current_user else None
        admin_ids = [u.id for u in User.query.filter_by(role='admin').all()]
        query = query.filter(
            db.or_(Product.created_by.is_(None), Product.created_by.in_(admin_ids), Product.created_by == uid)
        )
    if category:
        query = query.filter(Product.category.ilike(f'%{category}%'))

    # v2.6.0 高级过滤参数
    if category_id:
        from models import ProductCategory
        query = query.filter(Product.id.in_(
            db.session.query(ProductCategory.product_id).filter(ProductCategory.category_id == category_id)
        ))
    if manufacturer_id:
        query = query.filter(Product.manufacturer_id == manufacturer_id)
    if comm_method:
        query = query.filter(Product.id.in_(
            db.session.query(ProductCommMethod.product_id)
            .filter(ProductCommMethod.method_id.in_(comm_method))
            .group_by(ProductCommMethod.product_id)
            .having(func.count(distinct(ProductCommMethod.method_id)) == len(comm_method))
        ))
    if sensor_metric:
        from models import ProductSensorCapability
        query = query.filter(Product.id.in_(
            db.session.query(ProductSensorCapability.product_id)
            .filter(ProductSensorCapability.metric_id.in_(sensor_metric))
            .group_by(ProductSensorCapability.product_id)
            .having(func.count(distinct(ProductSensorCapability.metric_id)) == len(sensor_metric))
        ))
    if power_supply:
        query = query.filter(Product.id.in_(
            db.session.query(ProductPowerSupply.product_id)
            .filter(ProductPowerSupply.power_id.in_(power_supply))
            .group_by(ProductPowerSupply.product_id)
            .having(func.count(distinct(ProductPowerSupply.power_id)) == len(power_supply))
        ))

    # 拼音搜索：纯ASCII（无汉字）时启用
    is_pinyin = search and not re.search(r'[\u4e00-\u9fff]', search)
    if is_pinyin:
        q_lower = search.lower().strip()
        like = f'%{q_lower}%'
        query = query.filter(
            db.or_(Product.pinyin_search.like(like), Product.spec.ilike(like))
        )
    elif search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(Product.name.ilike(like), Product.spec.ilike(like),
                    Product.function_desc.ilike(like))
        )

    query = query.order_by(col.asc() if sort_order == 'asc' else col.desc())
    total = query.count()
    products = query.offset((page - 1) * per_page).limit(per_page).all()

    # 获取所有分类标签（支持逗号分隔的多标签）
    from sqlalchemy import text as _text
    raw_cats = [r[0] for r in db.session.execute(_text('SELECT category FROM products WHERE category IS NOT NULL')).all() if r[0]]
    cat_set = set()
    for c in raw_cats:
        for tag in c.split(','):
            tag = tag.strip()
            if tag:
                cat_set.add(tag)
    categories = sorted(cat_set)

    suppliers_list = sorted([r[0] for r in db.session.execute(_text('SELECT DISTINCT supplier FROM products WHERE supplier IS NOT NULL')).all() if r[0]])
    latest = db.session.query(func.max(Product.updated_at)).scalar()
    total_all = Product.query.count()

    # 预加载创建者用户名，避免 N+1 查询
    creator_ids = list(set(p.created_by for p in products if p.created_by))
    users_map = {}
    if creator_ids:
        creator_users = User.query.filter(User.id.in_(creator_ids)).all()
        users_map = {u.id: u.username for u in creator_users}

    # v2.6.0: 预加载分类/制造商/供应商/产品类型名称
    from models import DeviceCategory, Manufacturer, Supplier, DictProductType
    from models import DictCommMethod, DictPowerSupply
    cat_ids = [p.category_id for p in products if p.category_id]
    cat_map = {}
    if cat_ids:
        cats = DeviceCategory.query.filter(DeviceCategory.id.in_(cat_ids)).all()
        cat_map = {c.id: c.name for c in cats}
    mfr_ids = [p.manufacturer_id for p in products if p.manufacturer_id]
    mfr_map = {}
    if mfr_ids:
        mfrs = Manufacturer.query.filter(Manufacturer.id.in_(mfr_ids)).all()
        mfr_map = {m.id: m.name for m in mfrs}
    sup_ids = [p.supplier_id for p in products if p.supplier_id]
    sup_map = {}
    if sup_ids:
        sups = Supplier.query.filter(Supplier.id.in_(sup_ids)).all()
        sup_map = {s.id: s.name for s in sups}
    type_ids = [p.product_type_id for p in products if p.product_type_id]
    type_map = {}
    if type_ids:
        types = DictProductType.query.filter(DictProductType.id.in_(type_ids)).all()
        type_map = {t.id: t.name for t in types}

    # v2.6.0: 预加载 M2M 数据和主图
    product_ids = [p.id for p in products]
    comm_map = {}  # product_id → [{method_name, method_type, ...}, ...]
    power_map = {}  # product_id → [{power_name, power_category, ...}, ...]
    image_map = {}  # product_id → primary image url
    if product_ids:
        comm_rows = db.session.query(ProductCommMethod).filter(
            ProductCommMethod.product_id.in_(product_ids)
        ).all()
        method_ids = [r.method_id for r in comm_rows]
        dm_map = {}
        if method_ids:
            dms = DictCommMethod.query.filter(DictCommMethod.id.in_(method_ids)).all()
            dm_map = {d.id: d for d in dms}
        for r in comm_rows:
            dm = dm_map.get(r.method_id)
            comm_map.setdefault(r.product_id, []).append({
                'method_id': r.method_id,
                'method_name': dm.name if dm else '',
                'method_type': dm.method_type if dm else '',
                'details': r.details or '',
            })

        power_rows = db.session.query(ProductPowerSupply).filter(
            ProductPowerSupply.product_id.in_(product_ids)
        ).all()
        power_ids = [r.power_id for r in power_rows]
        dp_map = {}
        if power_ids:
            dps = DictPowerSupply.query.filter(DictPowerSupply.id.in_(power_ids)).all()
            dp_map = {d.id: d for d in dps}
        for r in power_rows:
            dp = dp_map.get(r.power_id)
            power_map.setdefault(r.product_id, []).append({
                'power_id': r.power_id,
                'power_name': dp.name if dp else '',
                'power_category': dp.supply_category if dp else '',
                'voltage_range': r.voltage_range or '',
                'battery_life': r.battery_life or '',
            })

        from models import ProductSensorCapability, DictSensorMetric
        sensor_rows = db.session.query(ProductSensorCapability).filter(
            ProductSensorCapability.product_id.in_(product_ids)
        ).all()
        sensor_map = {}
        if sensor_rows:
            metric_ids = [r.metric_id for r in sensor_rows]
            sm_dict = {}
            if metric_ids:
                sms = DictSensorMetric.query.filter(DictSensorMetric.id.in_(metric_ids)).all()
                sm_dict = {s.id: s.name for s in sms}
            for r in sensor_rows:
                sensor_map.setdefault(r.product_id, []).append({
                    'metric_id': r.metric_id,
                    'metric_name': sm_dict.get(r.metric_id, ''),
                    'measure_range': r.measure_range or '',
                })

        # Preload product_categories M2M
        from models import ProductCategory
        pc_rows = db.session.query(ProductCategory).filter(
            ProductCategory.product_id.in_(product_ids)
        ).all()
        cat_names_map = {}
        if pc_rows:
            all_cat_ids = {r.category_id for r in pc_rows}
            cat_name_lookup = {}
            if all_cat_ids:
                cat_objs = DeviceCategory.query.filter(DeviceCategory.id.in_(all_cat_ids)).all()
                cat_name_lookup = {c.id: c.name for c in cat_objs}
            for r in pc_rows:
                name = cat_name_lookup.get(r.category_id, '')
                if name:
                    cat_names_map.setdefault(r.product_id, []).append(name)

        primary_images = ProductImage.query.filter(
            ProductImage.product_id.in_(product_ids),
            ProductImage.is_primary == True
        ).order_by(ProductImage.sort_order).all()
        seen = set()
        for img in primary_images:
            if img.product_id not in seen:
                image_map[img.product_id] = img.url
                seen.add(img.product_id)

    def _thumb_url(url):
        """Convert image URL to thumbnail URL."""
        if not url or url.startswith('http'):
            return url
        if '_thumb.' in url:
            return url
        dot = url.rfind('.')
        if dot > 0:
            return url[:dot] + '_thumb' + url[dot:]
        return url

    products_json = []
    for p in products:
        p_dict = add_pinyin_field(p.to_dict(users_map=users_map))
        p_dict['category_name'] = cat_map.get(p.category_id, '')
        p_dict['manufacturer_name'] = mfr_map.get(p.manufacturer_id, '')
        p_dict['supplier_name'] = sup_map.get(p.supplier_id, '')
        p_dict['product_type_name'] = type_map.get(p.product_type_id, '')
        p_dict['comm_methods'] = comm_map.get(p.id, [])
        p_dict['power_supplies'] = power_map.get(p.id, [])
        p_dict['sensor_capabilities'] = sensor_map.get(p.id, [])
        p_dict['category_names'] = cat_names_map.get(p.id, [])
        img_url = p_dict.get('image_url', '') or image_map.get(p.id, '')
        p_dict['image_url'] = _thumb_url(img_url)
        p_dict['full_image_url'] = img_url  # 保留原图给其他用途
        if not is_admin:
            p_dict = _filter_fields_for_user(p_dict, is_admin)
        products_json.append(p_dict)

    return jsonify({
        'products': products_json,
        'total': total,
        'page': page,
        'per_page': per_page,
        'categories': sorted(categories),
        'suppliers': sorted(suppliers_list),
        'version': {'count': total_all, 'max_updated_at': latest.isoformat() if latest else None},
    })


@products_bp.route('/api/products', methods=['POST'])
def create_product():
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'error': '产品名称不能为空'}), 400

    name = str(data['name']).strip()
    if not name:
        return jsonify({'error': '产品名称不能为空'}), 400
    # Reject XSS vectors in product name
    xss_patterns = ['<script', '<img', 'onerror=', 'onclick=', 'onload=', 'javascript:']
    if any(p in name.lower() for p in xss_patterns):
        return jsonify({'error': '产品名称包含非法字符'}), 400
    # Enforce 20-char limit on product name
    if len(name) > 20:
        return jsonify({'error': '产品名称不能超过20个字'}), 400

    # 规格型号统一：spec 为主，同时填充 sku
    spec = data.get('spec', '')
    product = Product(
        name=name,
        sku=spec,
        category=data.get('category', ''),
        spec=spec,
        unit=data.get('unit', ''),
        price=round(float(data.get('price', 0)), 2),
        cost_price=round(float(data.get('cost_price', 0)), 2),
        supplier=data.get('supplier', ''),
        function_desc=data.get('function_desc', ''),
        remark=data.get('remark', ''),
        image_url=data.get('image_url', ''),
        created_by=g.current_user.id if hasattr(g, 'current_user') and g.current_user else None,
    )
    _store_image_blob(product, data)
    product.pinyin_search = _compute_pinyin_search(name, spec, data.get('category', ''), data.get('supplier', ''))
    # Auto-create product type from custom name
    if data.get('product_type_name'):
        from models import DictProductType
        pt = DictProductType.query.filter_by(name=data['product_type_name'].strip()).first()
        if not pt:
            pt = DictProductType(name=data['product_type_name'].strip())
            db.session.add(pt)
            db.session.flush()
        product.product_type_id = pt.id

    # Auto-create manufacturer from custom name
    if data.get('manufacturer_name'):
        from models import Manufacturer
        mfr = Manufacturer.query.filter_by(name=data['manufacturer_name'].strip()).first()
        if not mfr:
            mfr = Manufacturer(name=data['manufacturer_name'].strip())
            db.session.add(mfr)
            db.session.flush()
        product.manufacturer_id = mfr.id

    # Auto-match category string to device_category
    if 'category' in data and data['category']:
        primary_cat = data['category'].split(',')[0].strip()
        from models import DeviceCategory
        matched = DeviceCategory.query.filter_by(name=primary_cat, is_active=True).first()
        if matched:
            product.category_id = matched.id
        elif 'category_id' not in data or data['category_id'] is None:
            # Create new category if not found
            new_cat = DeviceCategory(name=primary_cat, is_active=True)
            db.session.add(new_cat)
            db.session.flush()
            product.category_id = new_cat.id

    # v2.6.0 new optional fields
    optional_new_fields = ['model', 'category_id', 'manufacturer_id', 'supplier_id',
                           'product_url', 'status', 'parent_id', 'unit', 'remark', 'product_type_id']
    for f in optional_new_fields:
        if f in data and data[f] is not None:
            setattr(product, f, data[f])
    for json_field in ['specs', 'urls', 'custom_fields']:
        if json_field in data:
            val = data[json_field]
            setattr(product, json_field, json.dumps(val, ensure_ascii=False) if val else None)
    # Save product_categories M2M
    if 'category_ids' in data:
        from models import ProductCategory
        ProductCategory.query.filter_by(product_id=product.id).delete()
        for cid in data['category_ids']:
            if cid:
                db.session.add(ProductCategory(product_id=product.id, category_id=cid))

    db.session.add(product)
    db.session.commit()
    return jsonify({'product': product.to_dict()}), 201


@products_bp.route('/api/products/<int:product_id>', methods=['GET'])
def get_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({'error': '产品不存在'}), 404
    p_dict = product.to_dict()
    # Load related names from FK tables
    from models import DeviceCategory, Manufacturer, Supplier, DictProductType
    from models import DictCommMethod, DictPowerSupply
    _cat = db.session.get(DeviceCategory, product.category_id) if product.category_id else None
    _mfr = db.session.get(Manufacturer, product.manufacturer_id) if product.manufacturer_id else None
    _sup = db.session.get(Supplier, product.supplier_id) if product.supplier_id else None
    if _cat:
        p_dict['category_name'] = _cat.name
    if _mfr:
        p_dict['manufacturer_name'] = _mfr.name
    if _sup:
        p_dict['supplier_name'] = _sup.name
    _type = db.session.get(DictProductType, product.product_type_id) if product.product_type_id else None
    if _type:
        p_dict['product_type_name'] = _type.name

    # Load M2M data
    comm_rows = ProductCommMethod.query.filter_by(product_id=product_id).all()
    method_ids = [r.method_id for r in comm_rows]
    dm_map = {}
    if method_ids:
        dms = DictCommMethod.query.filter(DictCommMethod.id.in_(method_ids)).all()
        dm_map = {d.id: d for d in dms}
    p_dict['comm_methods'] = [{
        'method_id': r.method_id,
        'method_name': dm_map[r.method_id].name if r.method_id in dm_map else '',
        'method_type': dm_map[r.method_id].method_type if r.method_id in dm_map else '',
        'dict_id': r.method_id,
        'detail': r.details or '',
    } for r in comm_rows]

    proto_rows = ProductCommProtocol.query.filter_by(product_id=product_id).all()
    proto_ids = [r.protocol_id for r in proto_rows]
    dp_map2 = {}
    if proto_ids:
        from models import DictCommProtocol
        dcp_list = DictCommProtocol.query.filter(DictCommProtocol.id.in_(proto_ids)).all()
        dp_map2 = {d.id: d.name for d in dcp_list}
    p_dict['comm_protocols'] = [{
        'protocol_id': r.protocol_id,
        'protocol_name': dp_map2.get(r.protocol_id, ''),
        'dict_id': r.protocol_id,
        'dict_name': dp_map2.get(r.protocol_id, ''),
        'direction': r.direction or 'both',
    } for r in proto_rows]

    power_rows = ProductPowerSupply.query.filter_by(product_id=product_id).all()
    power_ids = [r.power_id for r in power_rows]
    dp_map = {}
    if power_ids:
        dps = DictPowerSupply.query.filter(DictPowerSupply.id.in_(power_ids)).all()
        dp_map = {d.id: d for d in dps}
    p_dict['power_supplies'] = [{
        'power_id': r.power_id,
        'power_name': dp_map[r.power_id].name if r.power_id in dp_map else '',
        'power_category': dp_map[r.power_id].supply_category if r.power_id in dp_map else '',
        'dict_id': r.power_id,
        'voltage_range': r.voltage_range or '',
        'battery_life': r.battery_life or '',
    } for r in power_rows]

    # Load images from product_images table
    img_rows = ProductImage.query.filter_by(product_id=product_id).order_by(ProductImage.sort_order).all()
    p_dict['images'] = [img.to_dict() for img in img_rows]

    # Load category names from M2M
    from models import ProductCategory
    pc_rows = ProductCategory.query.filter_by(product_id=product_id).all()
    cat_ids = [r.category_id for r in pc_rows]
    if cat_ids:
        cat_objs = DeviceCategory.query.filter(DeviceCategory.id.in_(cat_ids)).all()
        p_dict['category_names'] = [c.name for c in cat_objs]
        p_dict['category_ids'] = [c.id for c in cat_objs]
        if cat_objs:
            p_dict['category_name'] = cat_objs[0].name
    else:
        p_dict['category_names'] = []
        p_dict['category_ids'] = []

    # Load hardware interfaces
    from models import ProductHardwareInterface
    hw_rows = ProductHardwareInterface.query.filter_by(product_id=product_id).order_by(ProductHardwareInterface.id).all()
    p_dict['hardware_interfaces'] = [{
        'id': hw.id,
        'interface_name': hw.interface_name,
        'quantity': hw.quantity or 1,
        'description': hw.description or '',
    } for hw in hw_rows]

    # Load sensor capabilities
    from models import ProductSensorCapability, DictSensorMetric
    sn_rows = ProductSensorCapability.query.filter_by(product_id=product_id).all()
    metric_ids = [s.metric_id for s in sn_rows]
    sm_map = {}
    if metric_ids:
        sms = DictSensorMetric.query.filter(DictSensorMetric.id.in_(metric_ids)).all()
        sm_map = {s.id: s for s in sms}
    p_dict['sensor_capabilities'] = [{
        'metric_id': s.metric_id,
        'metric_name': sm_map[s.metric_id].name if s.metric_id in sm_map else '',
        'unit': sm_map[s.metric_id].unit if s.metric_id in sm_map else '',
        'dict_id': s.metric_id,
        'measure_range': s.measure_range or '',
        'accuracy': s.accuracy or '',
        'resolution': s.resolution or '',
    } for s in sn_rows]

    is_admin = hasattr(g, 'current_user') and g.current_user and g.current_user.role == 'admin'
    if not is_admin:
        p_dict = _filter_fields_for_user(p_dict, is_admin)
    return jsonify({'product': p_dict})


@products_bp.route('/api/products/<int:product_id>/image', methods=['GET'])
def get_product_image(product_id):
    """返回产品图片二进制数据（支持 query param token 认证）"""
    # 不用 @require_auth，手动验证（img 标签无法设 header）
    from flask import request as _req
    token = _req.headers.get('Authorization', '').replace('Bearer ', '') or _req.args.get('token', '')
    if not token:
        return '', 401
    try:
        import jwt as _jwt
        from flask import current_app
        data = _jwt.decode(token, current_app.config['JWT_SECRET'], algorithms=['HS256'])
        from models import User
        user = db.session.get(User, data['user_id'])
        if not user or not user.is_active:
            return '', 403
    except Exception:
        return '', 401
    product = db.session.get(Product, product_id)
    if not product:
        return '', 404
    # 优先从 product_images 表获取主图 URL
    from models import ProductImage as PImg
    primary = PImg.query.filter_by(product_id=product_id, is_primary=True).first()
    img_url = primary.url if primary else (product.image_url or '')
    if not img_url:
        return '', 404
    if img_url.startswith('/uploads/'):
        fpath = BASE_DIR / img_url.lstrip('/')
        if fpath.exists():
            return send_file(str(fpath))
    # Fallback: 直接返回 BLOB 旧数据
    if product.image_data:
        return Response(product.image_data, mimetype=product.image_mime or 'image/jpeg')
    return '', 404


@products_bp.route('/api/products/<int:product_id>', methods=['PUT'])
def update_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({'error': '产品不存在'}), 404
    # 权限检查：管理员可编辑全部，普通用户只能编辑自己创建的
    is_admin = hasattr(g, 'current_user') and g.current_user and g.current_user.role == 'admin'
    if not is_admin and product.created_by != g.current_user.id:
        return jsonify({'error': '只能编辑自己创建的产品'}), 403
    data = request.get_json()
    for field in ['name', 'sku', 'category', 'spec', 'unit', 'supplier', 'function_desc', 'remark', 'image_url']:
        if field in data:
            val = data[field]
            if field == 'name':
                val = str(val).strip()
                if not val:
                    return jsonify({'error': '产品名称不能为空'}), 400
                xss_patterns = ['<script', '<img', 'onerror=', 'onclick=', 'onload=', 'javascript:']
                if any(p in val.lower() for p in xss_patterns):
                    return jsonify({'error': '产品名称包含非法字符'}), 400
                if len(val) > 20:
                    return jsonify({'error': '产品名称不能超过20个字'}), 400
            setattr(product, field, val)
    # 规格型号统一
    if product.spec and product.sku != product.spec:
        product.sku = product.spec
    if not product.spec and product.sku:
        product.spec = product.sku
    if 'price' in data:
        product.price = round(float(data['price']), 2)
    if 'cost_price' in data:
        product.cost_price = round(float(data['cost_price']), 2)
    # Auto-create product type from custom name
    if data.get('product_type_name'):
        from models import DictProductType
        pt = DictProductType.query.filter_by(name=data['product_type_name'].strip()).first()
        if not pt:
            pt = DictProductType(name=data['product_type_name'].strip())
            db.session.add(pt)
            db.session.flush()
        product.product_type_id = pt.id

    # Auto-create manufacturer from custom name
    if data.get('manufacturer_name'):
        from models import Manufacturer
        mfr = Manufacturer.query.filter_by(name=data['manufacturer_name'].strip()).first()
        if not mfr:
            mfr = Manufacturer(name=data['manufacturer_name'].strip())
            db.session.add(mfr)
            db.session.flush()
        product.manufacturer_id = mfr.id

    # Auto-match category string to device_category
    if 'category' in data and data['category']:
        primary_cat = data['category'].split(',')[0].strip()
        from models import DeviceCategory
        matched = DeviceCategory.query.filter_by(name=primary_cat, is_active=True).first()
        if matched:
            product.category_id = matched.id
        elif 'category_id' not in data or data['category_id'] is None:
            # Create new category if not found
            new_cat = DeviceCategory(name=primary_cat, is_active=True)
            db.session.add(new_cat)
            db.session.flush()
            product.category_id = new_cat.id

    # v2.6.0 new optional fields
    optional_new_fields = ['model', 'category_id', 'manufacturer_id', 'supplier_id',
                           'product_url', 'status', 'parent_id', 'unit', 'remark', 'product_type_id']
    for f in optional_new_fields:
        if f in data and data[f] is not None:
            setattr(product, f, data[f])
    for json_field in ['specs', 'urls', 'custom_fields']:
        if json_field in data:
            val = data[json_field]
            setattr(product, json_field, json.dumps(val, ensure_ascii=False) if val else None)
    _store_image_blob(product, data)
    product.pinyin_search = _compute_pinyin_search(product.name, product.spec or '', product.category or '', product.supplier or '')

    # Save product_categories M2M
    if 'category_ids' in data:
        from models import ProductCategory
        ProductCategory.query.filter_by(product_id=product.id).delete()
        for cid in data['category_ids']:
            if cid:
                db.session.add(ProductCategory(product_id=product.id, category_id=cid))

    db.session.commit()
    return jsonify({'product': product.to_dict()})


@products_bp.route('/api/products/<int:product_id>', methods=['DELETE'])
def delete_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({'error': '产品不存在'}), 404
    # 权限检查：管理员可删除全部，普通用户只能删除自己创建的
    is_admin = hasattr(g, 'current_user') and g.current_user and g.current_user.role == 'admin'
    if not is_admin and product.created_by != g.current_user.id:
        return jsonify({'error': '只能删除自己创建的产品'}), 403
    _cleanup_product_images(product_id)
    db.session.delete(product)
    db.session.commit()
    return jsonify({'message': '已删除'})


@products_bp.route('/api/products/batch-delete', methods=['POST'])
def batch_delete_products():
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请选择要删除的产品'}), 400
    is_admin = hasattr(g, 'current_user') and g.current_user and g.current_user.role == 'admin'
    if not is_admin:
        # 普通用户：只允许删除自己创建的产品，过滤掉非自己的
        uid = g.current_user.id
        own_ids = [p.id for p in Product.query.filter(
            Product.id.in_(ids), Product.created_by == uid
        ).all()]
        if not own_ids:
            return jsonify({'error': '没有可以删除的产品'}), 403
        Product.query.filter(Product.id.in_(own_ids)).delete(synchronize_session=False)
        db.session.commit()
        skipped = len(ids) - len(own_ids)
        msg = f'已删除 {len(own_ids)} 个产品'
        if skipped > 0:
            msg += f'（跳过 {skipped} 个非自己创建的产品）'
        return jsonify({'message': msg})
    Product.query.filter(Product.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'message': f'已删除 {len(ids)} 个产品'})


@products_bp.route('/api/products/version', methods=['GET'])
def products_version():
    count = Product.query.count()
    latest = db.session.query(func.max(Product.updated_at)).scalar()
    return jsonify({
        'count': count,
        'max_updated_at': latest.isoformat() if latest else None
    })


# ─── 图片OCR识别接口 ─────────────────────────────────────────

@products_bp.route('/api/products/ocr', methods=['POST'])
def ocr_image():
    """上传图片进行OCR识别，返回识别文本"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传图片文件'}), 400

    tmp_path = UPLOAD_DIR / f'_ocr_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
    try:
        # 保存临时文件
        file.save(str(tmp_path))
        size = os.path.getsize(tmp_path)
        if size > 5 * 1024 * 1024:
            return jsonify({'error': '图片不能超过5MB'}), 400

        # 调用OCR.space免费API
        import requests as http_req
        with open(tmp_path, 'rb') as fp:
            r = http_req.post(
                'https://api.ocr.space/parse/image',
                files={'file': fp},
                data={
                    'language': 'chs',
                    'isOverlayRequired': False,
                    'detectOrientation': True,
                    'scale': True,
                    'apikey': os.environ.get('OCR_SPACE_API_KEY', ''),
                },
                timeout=30,
            )

        if r.status_code != 200:
            return jsonify({'error': 'OCR服务暂时不可用'}), 502

        result = r.json()
        if result.get('OCRExitCode') == 1:
            text = result.get('ParsedResults', [{}])[0].get('ParsedText', '')
            return jsonify({'text': text.strip()})
        else:
            err = result.get('ErrorMessage', ['识别失败'])[0]
            return jsonify({'error': f'OCR识别失败: {err}'}), 400
    except Exception as e:
        return jsonify({'error': f'OCR处理失败: {str(e)}'}), 500
    finally:
        try: os.remove(tmp_path)
        except Exception: pass


@products_bp.route('/api/products/recognize', methods=['POST'])
def recognize_product():
    """智能识别粘贴内容（文字或图片），提取产品信息。
    每次只识别1个产品。
    图片用豆包 Vision；OCR/文字用 DeepSeek v4 Flash 解析（regex 降级）。
    请求体：{"text": "..."} 或上传 file 字段的图片
    返回：{"products": [...], "source": "...", "raw_text": "..."}
    """
    import time as _time
    _t0 = _time.time()
    _user_id = getattr(g, 'current_user', None)
    _user_id = _user_id.id if _user_id else 0

    data = request.get_json(silent=True) or {}
    uploaded_file = request.files.get('file')

    def _respond(product, source):
        raw = product.pop('_raw', '') if product else ''
        elapsed = _time.time() - _t0
        _log_ai_usage(user_id=_user_id, action='recognize', model=source, elapsed=elapsed, success=bool(product))
        return jsonify({'products': [product] if product else [], 'source': source, 'raw_text': raw[:3000]})

    text = None

    # 模式1: 图片文件上传 → 豆包 Vision
    if uploaded_file:
        tmp_path = UPLOAD_DIR / f'_smart_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
        try:
            uploaded_file.save(str(tmp_path))
            size = os.path.getsize(tmp_path)
            if size > 5 * 1024 * 1024:
                return jsonify({'error': '图片不能超过5MB'}), 400

            import base64
            with open(tmp_path, 'rb') as fp:
                image_b64 = base64.b64encode(fp.read()).decode('utf-8')

            product = doubao_vision_recognize(image_b64)
            if product:
                return _respond(product, 'doubao-vision')

            # 降级：OCR.space → DeepSeek 解析
            text = _ocr_fallback(str(tmp_path))
            if text:
                product = deepseek_parse_product(text)
                if product:
                    return _respond(product, 'deepseek-parse')
                # DeepSeek 失败 → regex 兜底
                product = smart_parse_product(text)
                if product:
                    return _respond(product, 'regex-parse')
            return jsonify({'products': [], 'error': '未能从图片中识别出产品信息，请检查图片清晰度'})
        except Exception as e:
            return jsonify({'error': f'图片处理失败: {str(e)}'}), 500
        finally:
            try: os.remove(tmp_path)
            except Exception: pass

    # 模式2: base64图片 → 豆包 Vision
    elif data.get('image'):
        try:
            import base64
            img_data = data['image']
            if ',' in img_data:
                img_data = img_data.split(',', 1)[1]

            product = doubao_vision_recognize(img_data, mime_type=data.get('mime_type', 'image/png'))
            if product:
                return _respond(product, 'doubao-vision')
            return jsonify({'products': [], 'error': '未能从图片中识别出产品信息，请检查图片清晰度'})
        except Exception as e:
            return jsonify({'error': f'图片处理失败: {str(e)}'}), 500

    # 模式3: 纯文本 → DeepSeek 解析（regex 降级）
    elif data.get('text', '').strip():
        text = data['text'].strip()
    else:
        return jsonify({'error': '请粘贴文字或图片'}), 400

    if not text:
        return jsonify({'products': [], 'error': '未能识别出文字内容'})

    # DeepSeek v4 Flash 优先
    product = deepseek_parse_product(text)
    if product:
        return _respond(product, 'deepseek-parse')

    # regex 降级
    product = smart_parse_product(text)
    if product:
        return _respond(product, 'regex-parse')
    return jsonify({'products': [], 'error': '未能从内容中识别出产品信息，请检查粘贴内容'})


# ─── 发票OCR → 成本价匹配 ─────────────────────────────────────

@products_bp.route('/api/products/ocr-costs', methods=['POST'])
@require_admin
def ocr_costs():
    """上传进货发票图片，OCR识别 + 自动匹配产品成本价"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传发票图片'}), 400

    tmp_path = UPLOAD_DIR / f'_ocr_cost_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
    try:
        file.save(str(tmp_path))
        if os.path.getsize(tmp_path) > 5 * 1024 * 1024:
            return jsonify({'error': '图片不能超过5MB'}), 400

        import requests as http_req
        with open(tmp_path, 'rb') as fp:
            r = http_req.post(
                'https://api.ocr.space/parse/image',
                files={'file': fp},
                data={'language': 'chs', 'isOverlayRequired': False, 'detectOrientation': True, 'scale': True, 'apikey': os.environ.get('OCR_SPACE_API_KEY', '')},
                timeout=30,
            )
        if r.status_code != 200:
            return jsonify({'error': 'OCR服务暂时不可用'}), 502

        result = r.json()
        if result.get('OCRExitCode') != 1:
            err = result.get('ErrorMessage', ['识别失败'])[0]
            return jsonify({'error': f'OCR失败: {err}'}), 400

        raw_text = result.get('ParsedResults', [{}])[0].get('ParsedText', '').strip()
        lines = [l.strip() for l in raw_text.split('\n') if l.strip()]

        # 解析每行：最后数字=成本价，前面=产品名
        products = Product.query.filter_by(is_active=True).all()
        matches = []
        import re
        for line in lines:
            # 找最后一个数字（可能带 ¥ 符号）
            m = re.findall(r'[¥￥]?\s*(\d+\.?\d*)\s*[元]?\s*$', line)
            if not m:
                continue
            try:
                cost_price = round(float(m[-1]), 2)
            except ValueError:
                continue
            # 去掉价格部分，剩余为产品描述
            name_part = re.sub(r'[¥￥]?\s*\d+\.?\d*\s*[元]?\s*$', '', line).strip()
            if not name_part or cost_price <= 0:
                continue

            # 模糊匹配产品
            candidates = []
            name_lower = name_part.lower()
            for p in products:
                score = 0
                p_name = (p.name or '').lower()
                p_spec = (p.spec or '').lower()
                p_supplier = (p.supplier or '').lower()
                if name_lower in p_name or p_name in name_lower:
                    score += 50
                if name_lower in p_spec or p_spec in name_lower:
                    score += 30
                if name_lower in p_supplier or p_supplier in name_lower:
                    score += 10
                # 检查空格分割的token匹配
                for token in name_lower.split():
                    if len(token) >= 2 and token in p_name:
                        score += 5
                    if len(token) >= 2 and token in p_spec:
                        score += 3
                if score > 0:
                    candidates.append({'id': p.id, 'name': p.name, 'spec': p.spec or '', 'cost_price': p.cost_price or 0, 'price': p.price or 0, 'supplier': p.supplier or '', 'score': score})
            candidates.sort(key=lambda x: -x['score'])
            matches.append({
                'line': line,
                'name_part': name_part,
                'cost_price': cost_price,
                'candidates': candidates[:3],
            })

        return jsonify({'raw_text': raw_text, 'matches': matches})

    except Exception as e:
        return jsonify({'error': f'处理失败: {str(e)}'}), 500
    finally:
        try: os.remove(tmp_path)
        except Exception: pass


@products_bp.route('/api/products/batch-costs', methods=['POST'])
@require_admin
def batch_update_costs():
    """批量更新产品成本价 {updates: [{id, cost_price}, ...]}"""
    data = request.get_json()
    if not data or not data.get('updates'):
        return jsonify({'error': '缺少更新数据'}), 400
    updated = 0
    for item in data['updates']:
        pid = item.get('id')
        cost = item.get('cost_price')
        if pid and cost is not None:
            product = db.session.get(Product, int(pid))
            if product:
                product.cost_price = round(float(cost), 2)
                updated += 1
    db.session.commit()
    return jsonify({'updated': updated, 'message': f'已更新{updated}个产品成本价'})


@products_bp.route('/api/products/<int:product_id>/toggle-active', methods=['PUT'])
@require_admin
def toggle_product_active(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({'error': '产品不存在'}), 404
    product.is_active = not product.is_active
    product.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'id': product.id, 'is_active': product.is_active})


@products_bp.route('/api/products/import', methods=['POST'])
@require_auth
def import_products():
    """从Excel导入产品 — 支持多Sheet、自动识别分类、提取嵌入图片"""
    import openpyxl

    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传Excel文件'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        imported, errors = _import_all_sheets(wb)
        db.session.commit()
        return jsonify({
            'message': f'成功导入 {imported} 个产品（共{len(wb.sheetnames)}个Sheet）',
            'imported': imported,
            'errors': errors,
        })
    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 400


@products_bp.route('/api/products/import-preview', methods=['POST'])
@require_auth
def import_preview():
    """上传Excel并返回解析预览数据（不导入）"""
    import openpyxl
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传Excel文件'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        preview_products = []
        total = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # Auto-detect header row (first row containing 名称, 产品, 型号 or price/单价)
            header_row_idx = 0
            header_keywords = ['名称', '型号', '产品名', 'name', 'product', '单价', 'price', '规格']
            for i, row in enumerate(rows):
                if not row:
                    continue
                cells = [str(c).strip() if c else '' for c in row]
                if any(any(kw in c.lower() if isinstance(c, str) else False for kw in header_keywords) for c in cells):
                    header_row_idx = i
                    break

            header_raw = [str(h).strip() if h else '' for h in rows[header_row_idx]]
            header = [h.lower() for h in header_raw]
            col_idx, _ = _parse_excel_header(ws, rows, header_row_idx)

            for row in rows[header_row_idx + 1:]:
                if not row:
                    continue
                try:
                    name_idx = col_idx.get('name', -1)
                    name = str(row[name_idx]).strip() if name_idx >= 0 and name_idx < len(row) and row[name_idx] else ''
                    if not name:
                        spec_idx = col_idx.get('spec', -1)
                        if spec_idx >= 0 and spec_idx < len(row) and row[spec_idx]:
                            name = str(row[spec_idx]).strip()
                    if not name:
                        continue

                    model_val = str(row[col_idx['model']]).strip() if col_idx.get('model', -1) >= 0 and col_idx.get('model', -1) < len(row) and row[col_idx['model']] else ''
                    if not model_val:
                        sku_idx = col_idx.get('sku', -1)
                        spec_idx = col_idx.get('spec', -1)
                        if sku_idx >= 0 and sku_idx < len(row) and row[sku_idx]:
                            model_val = str(row[sku_idx]).strip()
                        elif spec_idx >= 0 and spec_idx < len(row) and row[spec_idx]:
                            model_val = str(row[spec_idx]).strip()

                    sup_idx = col_idx.get('supplier', -1)
                    supplier_val = str(row[sup_idx]).strip() if sup_idx >= 0 and sup_idx < len(row) and row[sup_idx] else ''

                    mfr_idx = col_idx.get('manufacturer', -1)
                    mfr_val = str(row[mfr_idx]).strip() if mfr_idx >= 0 and mfr_idx < len(row) and row[mfr_idx] else supplier_val

                    # Check if product exists
                    from models import Product as Prod
                    existing = Prod.query.filter_by(name=name).first() or (Prod.query.filter_by(spec=model_val).first() if model_val else None)

                    # Smart classify product type
                    type_name = ''
                    text = f'{name} {model_val} {(str(row[col_idx["function_desc"]]) if col_idx.get("function_desc", -1) >= 0 and col_idx["function_desc"] < len(row) and row[col_idx["function_desc"]] else "")}'
                    type_rules = [
                        (r'网关|数传|DTU|RTU|采集器|gateway', '网关'),
                        (r'路由器|CPE|router', '路由器'),
                        (r'控制器|控制面板|集控|PLC|I/O|开关面板|插座|温控|窗帘|电机|执行器', '控制器'),
                        (r'电源|适配器|供电|配电', '电源'),
                        (r'支架|安装盒|线缆|天线|工具|附件|辅材|读卡器|打印机|磁力锁|按钮|蜂鸣器|机箱', '配件'),
                        (r'平板|交互屏|一体机|访客机|考勤|商显|标牌|工位屏|触控|摄像头|相机', '终端设备'),
                    ]
                    for pattern, tname in type_rules:
                        if re.search(pattern, text, re.IGNORECASE):
                            type_name = tname
                            break
                    if not type_name:
                        type_name = '传感器'  # default

                    preview_products.append({
                        'name': name,
                        'model': model_val,
                        'sku': model_val,
                        'category': sheet_name,
                        'product_type': type_name,
                        'manufacturer': mfr_val,
                        'supplier': supplier_val or mfr_val,
                        'price': _safe_number(row[col_idx['price']]) if col_idx.get('price', -1) >= 0 and col_idx['price'] < len(row) else 0,
                        'function_desc': str(row[col_idx['function_desc']]).strip() if col_idx.get('function_desc', -1) >= 0 and col_idx['function_desc'] < len(row) and row[col_idx['function_desc']] else '',
                        '_selected': not bool(existing),
                        '_status': 'exists' if existing else 'new',
                    })
                    total += 1
                except Exception:
                    continue

        return jsonify({'products': preview_products, 'total': total, 'sheets': len(wb.sheetnames)})
    except Exception as e:
        return jsonify({'error': f'解析失败: {str(e)}'}), 400


@products_bp.route('/api/products/import-confirm', methods=['POST'])
@require_auth
def import_confirm():
    """确认导入预览中选中的产品"""
    data = request.get_json()
    products_data = data.get('products', [])
    if not products_data:
        return jsonify({'error': '没有选择产品'}), 400

    imported = 0
    skipped = 0
    sheet_supplier_ref = ['']

    for item in products_data:
        name = (item.get('name') or '').strip()
        if not name:
            continue

        from models import Product as Prod
        existing = Prod.query.filter_by(name=name).first()
        if existing and not item.get('model'):
            existing.spec = item.get('model', '') or existing.spec
            existing.price = item.get('price', 0) or existing.price
            existing.function_desc = item.get('function_desc', '') or existing.function_desc
            existing.supplier = item.get('supplier', '') or existing.supplier
            existing.category = item.get('category', '') or existing.category
            existing.pinyin_search = _compute_pinyin_search(existing.name, existing.spec or '', existing.category or '', existing.supplier or '')
            db.session.add(existing)
            imported += 1
            continue

        spec_val = item.get('model', '') or item.get('sku', '')
        product = Product(
            name=name,
            sku=spec_val,
            spec=spec_val,
            category=item.get('category', ''),
            supplier=item.get('supplier', ''),
            function_desc=item.get('function_desc', ''),
            price=float(item.get('price', 0) or 0),
            created_by=g.current_user.id,
            remark='',
        )
        product.model = spec_val[:100]
        product.pinyin_search = _compute_pinyin_search(name, spec_val, item.get('category', ''), item.get('supplier', ''))
        product.category_id = _get_or_create_category(item.get('category', ''))
        mfr_name = item.get('manufacturer', '')
        if mfr_name:
            product.manufacturer_id = _get_or_create_manufacturer(mfr_name)
        sup_name = item.get('supplier', '')
        if sup_name:
            product.supplier_id = _get_or_create_supplier(sup_name)

        db.session.add(product)
        imported += 1

    db.session.commit()
    return jsonify({'imported': imported, 'skipped': skipped})


# ─── import_products 子函数 ────────────────────────────────────


def _get_or_create_category(name):
    """查找或创建设备分类，返回 category_id"""
    if not name or not name.strip():
        return None
    name = name.strip()
    from models import DeviceCategory
    cat = DeviceCategory.query.filter_by(name=name).first()
    if not cat:
        cat = DeviceCategory(name=name, level=1, sort_order=0)
        db.session.add(cat)
        db.session.flush()
    return cat.id


def _get_or_create_supplier(name):
    """查找或创建供应商，返回 supplier_id"""
    if not name or not name.strip():
        return None
    name = name.strip()
    from models import Supplier
    sup = Supplier.query.filter_by(name=name).first()
    if not sup:
        sup = Supplier(name=name)
        db.session.add(sup)
        db.session.flush()
    return sup.id


def _get_or_create_manufacturer(name):
    """查找或创建制造商，返回 manufacturer_id"""
    if not name or not name.strip():
        return None
    name = name.strip()
    from models import Manufacturer
    mfr = Manufacturer.query.filter_by(name=name).first()
    if not mfr:
        mfr = Manufacturer(name=name)
        db.session.add(mfr)
        db.session.flush()
    return mfr.id


_FIELD_MAP = {
    'name': ['产品名称', '名称', '品名', 'name', 'product'],
    'sku': ['编号', 'sku', '编码', '货号', '产品编号', '料号'],
    'spec': ['规格', '型号', '规格型号', 'spec', 'model', '功能/型号'],
    'unit': ['单位', 'unit'],
    'price': ['单价', '价格', '售价', '销售价', 'price', 'unit price'],
    'cost_price': ['成本价', '成本', '进价', '采购价', 'cost'],
    'supplier': ['供应商', '厂商', 'supplier'],
    'function_desc': ['功能描述'],
    'remark': ['备注', '说明', 'remark'],
    'image_url': ['图片', 'image', 'image_url', '产品图片'],
    'model': ['型号', '产品型号', 'model', '型号/规格'],
    'manufacturer': ['制造商', '品牌', 'manufacturer', '厂家', '生产商'],
    'product_url': ['产品链接', '官网链接', 'product_url', '链接'],
    'status': ['状态', 'status'],
}


def _find_col(header, names):
    """在表头中查找包含指定名称的列索引"""
    for i, h in enumerate(header):
        if not h:
            continue
        for n in names:
            if n in h or h in n:
                return i
    return -1


def _llm_parse_columns(headers, sample_rows):
    """使用LLM智能识别Excel列映射到数据库字段"""
    try:
        import urllib.request as _ur
        api_key = os.environ.get('DEEPSEEK_API_KEY', '')
        if not api_key:
            return None

        # Build prompt with headers and sample data
        header_list = '\n'.join(f'  Col{i}: "{h}"' for i, h in enumerate(headers) if h)
        sample_text = ''
        for ri, row in enumerate(sample_rows[:3]):
            row_text = ' | '.join(str(c)[:50] if c else '' for c in row)
            sample_text += f'\nRow {ri+1}: {row_text}'

        prompt = f"""你是一个产品数据库专家。请将以下Excel列映射到系统数据库字段。

可用字段: name(产品名称), model(型号), sku(SKU/编号), spec(规格), unit(单位), price(单价), cost_price(成本价), supplier(供应商), manufacturer(制造商/品牌), function_desc(功能描述), remark(内部备注), product_url(产品链接), status(状态), category(品类分类), product_type(产品类型如传感器/网关/控制器)

Excel表头:
{header_list}

数据样例:
{sample_text}

请分析每个表头和数据内容，返回JSON格式的列映射。格式: {{"字段名": 列索引}}
只返回JSON，不要其他文字。"""

        req = _ur.Request('https://api.deepseek.com/v1/chat/completions',
            data=json.dumps({
                'model': 'deepseek-chat',
                'messages': [{'role': 'user', 'content': prompt}],
                'max_tokens': 500,
                'temperature': 0.1,
            }).encode(),
            headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'},
            method='POST')
        with _ur.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            content = data['choices'][0]['message']['content'].strip()
            # Extract JSON from response
            m = re.search(r'\{[^}]+\}', content)
            if m:
                mapping = json.loads(m.group(0))
                # Convert to col_idx format
                result = {}
                for field, col_num in mapping.items():
                    field = field.strip().lower()
                    col_num = int(col_num) if isinstance(col_num, str) else col_num
                    if field in _FIELD_MAP or field in ['category', 'product_type', 'manufacturer']:
                        result[field] = col_num
                return result
            return None
    except Exception as e:
        print(f'[LLM parse] Error: {e}')
        return None


def _parse_excel_header(ws, rows, header_row_idx):
    """解析Excel表头，返回列索引映射 + 嵌入图片索引"""
    header_raw = [str(h).strip() if h else '' for h in rows[header_row_idx]]
    header = [h.lower() for h in header_raw]

    # Try LLM first for intelligent column mapping
    col_idx = {}
    sample_start = header_row_idx + 1
    samples = rows[sample_start:sample_start + 5] if len(rows) > sample_start else []
    llm_result = _llm_parse_columns(header_raw, samples)
    if llm_result:
        col_idx = llm_result
        print(f'[Import] LLM mapped {len(col_idx)} columns')
    else:
        # Fallback: hardcoded mapping
        for key, names in _FIELD_MAP.items():
            col_idx[key] = _find_col(header, names)

    # 智能解析：备注 vs 内部备注 vs 图片
    inner_remark_col = _find_col(header, ['内部备注'])
    if inner_remark_col >= 0 and col_idx.get('remark', -1) >= 0:
        if col_idx.get('image_url', -1) < 0:
            col_idx['image_url'] = col_idx['remark']
        col_idx['remark'] = inner_remark_col
    elif col_idx.get('image_url', -1) < 0 and col_idx.get('remark', -1) >= 0:
        pass  # 仅有一个备注列，保持为 remark

    # 构建嵌入图片索引
    image_map = {}
    if hasattr(ws, '_images'):
        for img in ws._images:
            try:
                anc = img.anchor
                if hasattr(anc, '_from'):
                    image_map[(anc._from.col, anc._from.row)] = img
            except Exception:
                pass

    return col_idx, image_map


def _detect_supplier_col(rows, header_row_idx):
    """供应商列回退：扫描数据行定位常见位置"""
    data_start2 = header_row_idx + 1
    for cc in [11, 12, 13]:
        if cc >= len(rows[header_row_idx]):
            continue
        sample_count = 0
        for dr in rows[data_start2:data_start2 + 10]:
            if dr and cc < len(dr) and dr[cc] and str(dr[cc]).strip():
                sample_count += 1
        if sample_count >= 2:
            return cc
    return -1


def _extract_embedded_image(emb_img):
    """从Excel嵌入图片提取并保存，返回 image_url 路径或空字符串"""
    try:
        img_bytes = emb_img._data()
        ext = (emb_img.format or 'png').lower()
        if ext == 'jpeg':
            ext = 'jpg'
        fname = f'prod_{int(datetime.now().timestamp())}_{random.randint(1000, 9999)}.{ext}'
        img_dir = UPLOAD_DIR / 'images'
        img_dir.mkdir(parents=True, exist_ok=True)
        save_path = img_dir / fname
        save_path.write_bytes(img_bytes)
        _, compressed_fname = compress_image_if_needed(str(save_path))
        return f'/uploads/images/{compressed_fname}'
    except Exception:
        return ''


def _process_import_row(row, row_idx, col_idx, image_map, sheet_name, sheet_supplier_ref, rows=None, header_row_idx=0):
    """处理单行数据，返回 (Product对象, error_string_or_None)"""
    name_idx = col_idx['name']
    name = str(row[name_idx]).strip() if name_idx >= 0 and name_idx < len(row) and row[name_idx] else ''
    if not name:
        spec_idx = col_idx.get('spec', -1)
        if spec_idx >= 0 and spec_idx < len(row) and row[spec_idx]:
            name = str(row[spec_idx]).strip()
    if not name:
        return None, None

    sup_val = str(row[col_idx['supplier']]).strip() if col_idx.get('supplier', -1) >= 0 and col_idx['supplier'] < len(row) and row[col_idx['supplier']] else ''
    if not sup_val and sheet_supplier_ref[0]:
        sup_val = sheet_supplier_ref[0]
    else:
        sheet_supplier_ref[0] = sup_val

    sku_val = str(row[col_idx['sku']]).strip() if col_idx.get('sku', -1) >= 0 and col_idx['sku'] < len(row) and row[col_idx['sku']] else ''
    spec_val = str(row[col_idx['spec']]).strip() if col_idx.get('spec', -1) >= 0 and col_idx['spec'] < len(row) and row[col_idx['spec']] else ''
    if spec_val:
        sku_val = spec_val
    else:
        spec_val = sku_val

    product = Product(
        name=name,
        category=sheet_name,
        sku=sku_val or spec_val,
        spec=spec_val,
        unit=str(row[col_idx['unit']]).strip() if col_idx.get('unit', -1) >= 0 and col_idx['unit'] < len(row) and row[col_idx['unit']] else '',
        price=_safe_number(row[col_idx['price']]) if col_idx.get('price', -1) >= 0 and col_idx['price'] < len(row) else 0,
        cost_price=_safe_number(row[col_idx['cost_price']]) if col_idx.get('cost_price', -1) >= 0 and col_idx['cost_price'] < len(row) else 0,
        supplier=sup_val,
        function_desc=str(row[col_idx['function_desc']]).strip() if col_idx.get('function_desc', -1) >= 0 and col_idx['function_desc'] < len(row) and row[col_idx['function_desc']] else '',
        remark=str(row[col_idx['remark']]).strip() if col_idx.get('remark', -1) >= 0 and col_idx['remark'] < len(row) and row[col_idx['remark']] else '',
        created_by=g.current_user.id if hasattr(g, 'current_user') and g.current_user else None,
    )

    # v2.6.0: populate new columns
    # model from dedicated column or spec
    model_idx = col_idx.get('model', -1)
    if model_idx >= 0 and model_idx < len(row) and row[model_idx]:
        product.model = str(row[model_idx]).strip()
    elif spec_val:
        product.model = spec_val[:100]

    # category_id from sheet name (auto-create if missing)
    product.category_id = _get_or_create_category(sheet_name)

    # supplier_id (auto-create if missing)
    if sup_val:
        product.supplier_id = _get_or_create_supplier(sup_val)

    # manufacturer_id (auto-create if missing)
    mfr_idx = col_idx.get('manufacturer', -1)
    if mfr_idx >= 0 and mfr_idx < len(row) and row[mfr_idx]:
        mfr_name = str(row[mfr_idx]).strip()
        product.manufacturer_id = _get_or_create_manufacturer(mfr_name)

    # product_type (auto-create if missing)
    type_idx = col_idx.get('product_type', -1)
    if type_idx >= 0 and type_idx < len(row) and row[type_idx]:
        type_name = str(row[type_idx]).strip()
        from models import DictProductType
        pt = DictProductType.query.filter_by(name=type_name).first()
        if not pt:
            pt = DictProductType(name=type_name)
            db.session.add(pt)
            db.session.flush()
        product.product_type_id = pt.id

    # category from explicit column (overrides sheet name)
    cat_idx = col_idx.get('category', -1)
    if cat_idx >= 0 and cat_idx < len(row) and row[cat_idx]:
        cat_name = str(row[cat_idx]).strip()
        product.category_id = _get_or_create_category(cat_name)

    # product_url
    url_idx = col_idx.get('product_url', -1)
    if url_idx >= 0 and url_idx < len(row) and row[url_idx]:
        product.product_url = str(row[url_idx]).strip()[:500]

    # status
    status_idx = col_idx.get('status', -1)
    if status_idx >= 0 and status_idx < len(row) and row[status_idx]:
        status_val = str(row[status_idx]).strip().lower()
        if status_val in ('active', 'archived', 'discontinued', 'planned', '在售', '停售', '规划中'):
            status_map = {'在售': 'active', '停售': 'discontinued', '规划中': 'planned'}
            product.status = status_map.get(status_val, status_val)

    # specs JSON: store any unrecognized columns as custom specs
    extra_specs = {}
    known_indices = set(v for v in col_idx.values() if v >= 0)
    known_indices.add(col_idx.get('image_url', -1))
    if rows is not None:
        for i, cell in enumerate(row):
            if i not in known_indices and cell is not None and str(cell).strip():
                header = rows[header_row_idx][i] if header_row_idx < len(rows) and i < len(rows[header_row_idx]) else ''
                key = str(header).strip() if header else f'_col_{i}'
                if key and key not in ('None', ''):
                    extra_specs[key] = str(cell).strip()
    if extra_specs:
        product.specs = json.dumps(extra_specs, ensure_ascii=False)

    # 提取图片：嵌入图片优先，URL 文本次之
    if col_idx.get('image_url', -1) >= 0:
        img_col_0 = col_idx['image_url']
        emb_img = image_map.get((img_col_0, row_idx - 1))
        if emb_img is not None:
            product.image_url = _extract_embedded_image(emb_img)
        if not product.image_url and img_col_0 < len(row) and row[img_col_0]:
            txt = str(row[img_col_0]).strip()
            if txt and (txt.startswith('http') or txt.startswith('/uploads/')):
                product.image_url = txt[:500]

    _store_image_blob(product, {'image_url': product.image_url or ''})
    return product, None


def _import_all_sheets(wb):
    """遍历所有Sheet导入产品，返回 (imported_count, errors_list)"""
    imported = 0
    errors = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        # 检测表头行
        first_nonempty = [v for v in rows[0] if v is not None and str(v).strip()]
        second_nonempty = [v for v in rows[1] if v is not None and str(v).strip()] if len(rows) > 1 else []
        header_row_idx = 1 if len(first_nonempty) <= 2 and len(second_nonempty) >= 3 else 0

        col_idx, image_map = _parse_excel_header(ws, rows, header_row_idx)
        if col_idx.get('name', -1) < 0:
            continue

        if col_idx.get('supplier', -1) < 0:
            col_idx['supplier'] = _detect_supplier_col(rows, header_row_idx)

        sheet_supplier_ref = ['']  # 用list以便在_process_import_row中修改
        data_start = header_row_idx + 1
        for row_idx, row in enumerate(rows[data_start:], data_start + 1):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            first_col = str(row[0]).strip().lower() if row[0] else ''
            if first_col in ('小计', '合计', '总计', 'subtotal', 'total', '注', '备注'):
                continue
            try:
                product, err = _process_import_row(row, row_idx, col_idx, image_map, sheet_name, sheet_supplier_ref, rows, header_row_idx)
                if product:
                    db.session.add(product)
                    imported += 1
            except Exception as e:
                errors.append(f'[{sheet_name}] 第{row_idx}行: {str(e)}')

    return imported, errors


@products_bp.route('/api/products/export-all', methods=['GET'])
@require_auth
@require_auth
def export_all_products():
    """导出全部产品为 Excel（按模板格式，管理员增加创建者列）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    is_admin = g.current_user.role == 'admin'
    uid = g.current_user.id

    # 权限过滤：普通用户只导出自己创建的产品
    query = Product.query
    if not is_admin:
        query = query.filter(Product.created_by == uid)
    products = query.order_by(Product.id.asc()).all()

    wb = openpyxl.Workbook()

    # Preload category names
    cat_ids = set(p.category_id for p in products if p.category_id)
    cat_map = {}
    if cat_ids:
        from models import DeviceCategory
        cats = DeviceCategory.query.filter(DeviceCategory.id.in_(cat_ids)).all()
        cat_map = {c.id: c.name for c in cats}

    # Preload manufacturer names + user names
    mfr_ids = set(p.manufacturer_id for p in products if p.manufacturer_id)
    mfr_map = {}
    if mfr_ids:
        from models import Manufacturer
        mfrs = Manufacturer.query.filter(Manufacturer.id.in_(mfr_ids)).all()
        mfr_map = {m.id: m for m in mfrs}
    user_ids = set(p.created_by for p in products if p.created_by)
    user_map = {}
    if user_ids:
        from models import User
        users = User.query.filter(User.id.in_(user_ids)).all()
        user_map = {u.id: u.username for u in users}

    # 按 category 分 Sheet，未分类放「未分类」
    sheet_map = {}
    for p in products:
        # Use new category name if available, else old string, else '未分类'
        cat_name = cat_map.get(p.category_id) if p.category_id else None
        if not cat_name:
            cat_name = (p.category or '').strip() or '未分类'
        if cat_name not in sheet_map:
            sheet_map[cat_name] = []
        sheet_map[cat_name].append(p)

    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(vertical='center', wrap_text=True)

    for sheet_name, prods in sheet_map.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        headers = ['产品名称', '型号', '规格型号', '功能描述', '备注', '供应商', '制造商', '单价', '成本价', '单位', '状态']
        if is_admin:
            headers.append('创建者')
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = center_align
        widths = [20, 15, 20, 30, 20, 15, 15, 10, 10, 8, 10]
        if is_admin:
            widths.append(10)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for p in prods:
            # Get manufacturer name
            mfr_name = ''
            if p.manufacturer_id and p.manufacturer_id in mfr_map:
                mfr_name = mfr_map[p.manufacturer_id].name if hasattr(mfr_map[p.manufacturer_id], 'name') else ''

            row_data = [
                p.name or '',
                p.model or '',
                p.spec or '',
                p.function_desc or '',
                p.remark or '',
                p.supplier or '',  # old string fallback
                mfr_name,
                p.price or 0,
                p.cost_price or 0,
                p.unit or '',
                p.status or 'active',
            ]
            if is_admin:
                row_data.append(user_map.get(p.created_by, ''))
            ws.append(row_data)

    # 删除默认空 Sheet
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = 'products_export.xlsx' if is_admin else 'my_products_export.xlsx'
    return send_file(output, download_name=fname, as_attachment=True)


@products_bp.route('/api/products/export-template', methods=['GET'])
def export_product_template():
    """下载原始报价规格库模板（包含所有分类Sheet）"""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    template_path = BASE_DIR / 'template.xlsx'
    if template_path.exists():
        return send_file(str(template_path), download_name='硬件报价规格库（成本）.xlsx', as_attachment=True)
    # 兜底：生成简易模板
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '产品模板'
    headers = ['序号', '名称', '规格型号', '功能/型号', '功能描述', '单价', '数量', '合计', '折扣率', '成交价', '备注', '供应商', '供应商型号', '成本', '指导价', '最低零售价', '备注']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')
    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 15
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='产品导入模板.xlsx', as_attachment=True)


# ─── 独立前缀路由（不在 /api/products 下） ────────────────────

@products_bp.route('/api/upload/image', methods=['POST'])
def upload_image():
    """上传产品图片"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请选择图片文件'}), 400

    # 校验文件类型
    allowed = {'image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/bmp'}
    if file.content_type not in allowed and not file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')):
        return jsonify({'error': '仅支持 JPG/PNG/GIF/WebP/BMP 格式'}), 400

    # 生成文件名
    ext = os.path.splitext(file.filename)[1] if '.' in file.filename else '.jpg'
    fname = f'prod_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(1000,9999)}{ext}'
    save_dir = UPLOAD_DIR / 'images'
    save_dir.mkdir(parents=True, exist_ok=True)
    filepath = save_dir / fname
    file.save(str(filepath))

    # 限制文件大小
    size = os.path.getsize(filepath)
    if size > 10 * 1024 * 1024:
        os.remove(filepath)
        return jsonify({'error': '图片不能超过10MB'}), 400

    # 压缩到 100KB 以内
    compressed_path, compressed_fname = compress_image_if_needed(str(filepath))
    if compressed_fname != fname:
        fname = compressed_fname

    # 返回相对URL
    image_url = f'/uploads/images/{fname}'
    return jsonify({'url': image_url, 'filename': fname})


@products_bp.route('/api/download-image', methods=['POST'])
def download_image():
    """从URL下载图片并保存到本地（防SSRF）"""
    import ipaddress
    import socket
    import urllib.request
    import base64

    data = request.get_json()
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'error': '请提供图片URL'}), 400
    # 只允许 http/https
    if not url.startswith(('http://', 'https://')):
        return jsonify({'error': '仅支持 http/https 链接'}), 400

    # ── SSRF 防护：解析域名，拒绝内网/保留IP ──
    try:
        from urllib.parse import urlparse
        hostname = urlparse(url).hostname
        if not hostname:
            return jsonify({'error': 'URL格式无效'}), 400
        resolved_ip = socket.getaddrinfo(hostname, None, socket.AF_UNSPEC, socket.SOCK_STREAM)
        for family, _, _, _, sockaddr in resolved_ip:
            ip = ipaddress.ip_address(sockaddr[0])
            if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast:
                return jsonify({'error': '不允许访问内网或保留地址'}), 400
    except socket.gaierror:
        return jsonify({'error': '域名解析失败'}), 400
    except Exception:
        return jsonify({'error': 'URL校验失败'}), 400

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            content = resp.read()
            content_type = resp.headers.get('Content-Type', '')
    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 400

    # 校验类型
    allowed_types = {'image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/bmp'}
    if content_type.split(';')[0].strip() not in allowed_types:
        return jsonify({'error': '仅支持 JPG/PNG/GIF/WebP/BMP 格式'}), 400

    # 推断扩展名
    ext_map = {'image/jpeg': '.jpg', 'image/png': '.png', 'image/gif': '.gif',
               'image/webp': '.webp', 'image/bmp': '.bmp'}
    ext = ext_map.get(content_type.split(';')[0].strip(), '.jpg')
    # 也检查 URL 扩展名
    url_path = url.split('?')[0]
    if '.' in url_path.rsplit('/', 1)[-1]:
        url_ext = os.path.splitext(url_path.rsplit('/', 1)[-1])[1].lower()
        if url_ext in {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}:
            ext = url_ext

    fname = f'prod_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(1000,9999)}{ext}'
    save_dir = UPLOAD_DIR / 'images'
    save_dir.mkdir(parents=True, exist_ok=True)
    filepath = save_dir / fname
    with open(filepath, 'wb') as f:
        f.write(content)

    # 压缩到 100KB 以内
    compressed_path, compressed_fname = compress_image_if_needed(str(filepath))
    if compressed_fname != fname:
        fname = compressed_fname

    image_url = f'/uploads/images/{fname}'
    # 读取压缩后的图片返回 base64，方便前端直接存入 BLOB
    import base64
    with open(save_dir / fname, 'rb') as f:
        img_b64 = base64.b64encode(f.read()).decode('utf-8')
    return jsonify({'url': image_url, 'filename': fname, 'image_data': img_b64, 'image_mime': 'image/jpeg'})
